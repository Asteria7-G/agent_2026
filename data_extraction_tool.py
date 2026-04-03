import time
import json
import re
import os
import base64
import requests
import fitz  # PyMuPDF
import pdfplumber
import pandas as pd
from tqdm import tqdm
from openai import AzureOpenAI
from openpyxl.styles import Font
from typing import Optional
from io import BytesIO
from nebula3.common.ttypes import ErrorCode
from nebula3.Config import Config
from nebula3.gclient.net import ConnectionPool
from nebula3.gclient.net.SessionPool import SessionPool
from webscoket_connect import websocket_client
import asyncio
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import cv2
import numpy as np


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

config = Config() # 定义一个配置
config.max_connection_pool_size = 10 # 设置最大连接数
connection_pool = ConnectionPool() # 初始化连接池
# 如果给定的服务器是ok的，返回true，否则返回false
ok = connection_pool.init([('10.8.140.25', 9669)], config)
print(ok)

azure_client = AzureOpenAI(
    azure_endpoint = "https://openaichatgpt-me-cn.openai.azure.com/",
    api_key = "a72b7770afac45d6ba000394ddde7151",
    api_version = "2025-01-01-preview"
)

###################################################################################################

                     #                   DIGITAL AGE 2025               #

###################################################################################################

def get_agent_reply(five_word):
    message_text = [
        {"role": "system",
         "content": "你是一个制造业数字化技术以及业务专家。"},
        {
            "role": "user",
            "content": (
                "我们现在正在进行制造业数字化技术以及业务讨论的沙龙，根据我输入的词语：%s，"
                "总结出一个高大上的话语（不超过50个词）,"
                "总结需要一份中文版和一份对照的英文翻译版。"
                "回复请返回一个string，里面严格遵守如下这样的排版："
                "'【中文版本: 生成的中文版本话语】 【English Verison: 生成的英文版本话语】’"
            ) % json.dumps(five_word)
        }
    ]

    response = azure_client.chat.completions.create(
        model="gpt-4o",
        messages=message_text,
        temperature=1,
        stream=False   # <-- VERY IMPORTANT
    )

    # 直接取一次性返回的文本，不需要累积
    reply = response.choices[0].message.content
    return reply.strip()





###################################################################################################

                     #                   NEBULA DATA                #

###################################################################################################

def get_node_info(tag):
    result_list = []
    with connection_pool.session_context('root', 'nebula') as session:
        session.execute('USE digitalSE_V5')
        result = session.execute(f'MATCH (v:{tag}) RETURN v')
        print(result)  # 打印查询结果
        for row in result:
            result_dict = {}
            node_n = row.values()[0]
            if node_n:
                node_obj = node_n.as_node()
                vid = node_obj.get_id()  # 获取 VID
                result_dict["id"] = vid
                props = dict(zip(node_obj.prop_names(tag), node_obj.prop_values(tag)))
                for prop_key, prop_values in props.items():
                    result_dict[prop_key] = prop_values
                for key in result_dict:
                    result_dict[key] = str(result_dict[key])[1:-1]
                result_list.append(result_dict)
    # result_list = str(result_list).replace("'", '"')
    # result_list = json.loads(result_list)

    return result_list

def get_relationship_info(source_tag, destination_tag):
    result_list = []
    with connection_pool.session_context('root', 'nebula') as session:
        session.execute('USE digitalSE_V5')
        result = session.execute(f'MATCH (v:{source_tag})-[e]->(n:{destination_tag}) RETURN v,e,n')
        print(result)  # 打印查询结果
        for row in result:
            result_dict = {}
            relationship = row.values()[1]
            print(f"RELATIONSHIP: {relationship}")
            if relationship:
                relationship_obj = relationship.as_relationship()
                # 获取 source node 和 destination node 的 VID
                relationship_type = relationship_obj.edge_name()
                source_vid = relationship_obj.start_vertex_id()
                destination_vid = relationship_obj.end_vertex_id()
                # 获取边的属性（字典形式）
                relationship_properties = relationship_obj.properties()

                # 存入字典
                result_dict["relationship_type"] = relationship_type
                result_dict["source_vid"] = source_vid
                result_dict["destination_vid"] = destination_vid
                result_dict["relationship_properties"] = relationship_properties
                for key in result_dict:
                    if key != "relationship_properties":
                        result_dict[key] = str(result_dict[key])[1:-1]
                    else:
                        for sub_key in result_dict[key]:
                            result_dict[key][sub_key] = str(result_dict[key][sub_key])[1:-1]
                result_list.append(result_dict)
    # result_list = str(result_list).replace("'", '"')
    # result_list = json.loads(result_list)
    return result_list

def get_node_info_from_customer_req(tag):
    result_list = []
    with connection_pool.session_context('root', 'nebula') as session:
        session.execute('USE digitalSE_customer_req')
        result = session.execute(f'MATCH (v:{tag}) RETURN v')
        print(result)  # 打印查询结果
        for row in result:
            result_dict = {}
            node_n = row.values()[0]
            if node_n:
                node_obj = node_n.as_node()
                vid = node_obj.get_id()  # 获取 VID
                result_dict["id"] = vid
                props = dict(zip(node_obj.prop_names(tag), node_obj.prop_values(tag)))
                for prop_key, prop_values in props.items():
                    result_dict[prop_key] = prop_values
                for key in result_dict:
                    result_dict[key] = str(result_dict[key])[1:-1]
                result_list.append(result_dict)
    # result_list = str(result_list).replace("'", '"')
    # result_list = json.loads(result_list)

    return result_list

def get_relationship_info_from_customer_req(source_tag, destination_tag):
    result_list = []
    with connection_pool.session_context('root', 'nebula') as session:
        session.execute('USE digitalSE_customer_req')
        result = session.execute(f'MATCH (v:{source_tag})-[e]->(n:{destination_tag}) RETURN v,e,n')
        print(result)  # 打印查询结果
        for row in result:
            result_dict = {}
            relationship = row.values()[1]
            print(f"RELATIONSHIP: {relationship}")
            if relationship:
                relationship_obj = relationship.as_relationship()
                # 获取 source node 和 destination node 的 VID
                relationship_type = relationship_obj.edge_name()
                source_vid = relationship_obj.start_vertex_id()
                destination_vid = relationship_obj.end_vertex_id()
                # 获取边的属性（字典形式）
                relationship_properties = relationship_obj.properties()

                # 存入字典
                result_dict["relationship_type"] = relationship_type
                result_dict["source_vid"] = source_vid
                result_dict["destination_vid"] = destination_vid
                result_dict["relationship_properties"] = relationship_properties
                for key in result_dict:
                    if key != "relationship_properties":
                        result_dict[key] = str(result_dict[key])[1:-1]
                    else:
                        for sub_key in result_dict[key]:
                            result_dict[key][sub_key] = str(result_dict[key][sub_key])[1:-1]
                result_list.append(result_dict)
    # result_list = str(result_list).replace("'", '"')
    # result_list = json.loads(result_list)
    return result_list


# for module to reference doc content
def reference_test_operation_step_and_check(module_vid_list):
    module_pn_dict = {}
    with connection_pool.session_context('root', 'nebula') as session:
        session.execute('USE digitalSE_V5')
        for module_vid in module_vid_list:
            result = session.execute(f'MATCH (p:PN)-[e]->(m:MODULE) WHERE id(m) == "{module_vid}" RETURN id(p);')
            pn_vid_list = []
            for row in result:
                pn_vid_list.append(str(row)[1:-1])
            module_pn_dict[module_vid] = pn_vid_list

    def find_common_pn(modules):
        # 获取第一个模块的pn列表
        first_key = list(modules.keys())[0]
        common_pn = set(modules[first_key])

        # 遍历所有模块的pn列表，并进行交集操作
        for module, pn_list in modules.items():
            common_pn &= set(pn_list)

        # 如果有共同的pn，返回它们，否则返回False
        return list(common_pn) if common_pn else False

    common_pn_result = find_common_pn(module_pn_dict)
    if common_pn_result:
        with connection_pool.session_context('root', 'nebula') as session:
            session.execute('USE digitalSE_V5')
            query = f'MATCH (p:PN)-[e1]->(d:PRV_DOCUMENT)-[e2]-(s:TEST_OPERATION_STEP) WHERE id(p) == "{common_pn_result[0]}" RETURN e2,s,toInteger(e2.`level_1_sorting `) AS l1, toInteger(e2.`level_2_sorting`) AS l2 ORDER BY l1,l2;'
            result = session.execute(query)
            #         print(result)
            result_list = []
            for row in result:
                step = row.values()[1]
                if step:
                    step_obj = step.as_node()
                    step_properties = step_obj.properties()
                    method_name = step_properties['parent_test_method_name']
                    sub_method_name = step_properties['subparent_method_name']
                    test_step = step_properties['description']
                #                 print(method_name,sub_method_name,test_step)
                relationship = row.values()[0]
                if relationship:
                    relationship_obj = relationship.as_relationship()
                    relationship_properties = relationship_obj.properties()
                    relationship_properties_level_2_sorting = relationship_properties["level_2_sorting"]
                    relationship_properties_parameter = relationship_properties["parameter"]

                #                 print(f"parameter:{relationship_properties_parameter}")

                first_tag = False
                for first_result_dict in result_list:
                    if first_result_dict["method_name"] == str(method_name)[1:-1]:
                        first_tag = True
                        second_tag = False
                        for second_dict in first_result_dict["test_step_sub_method_list"]:
                            if second_dict["sub_method_name"] == str(sub_method_name)[1:-1]:
                                second_dict["test_step_list"].append(
                                    {"test_step_number": f"step{str(relationship_properties_level_2_sorting)[1:-1]}",
                                     "test_step_content": str(test_step)[1:-1],
                                     "test_step_parameter": str(relationship_properties_parameter)[1:-1]})
                                second_tag = True
                                break
                        if second_tag == False:
                            first_result_dict["test_step_sub_method_list"].append({"sub_method_name": str(sub_method_name)[1:-1],
                                                                         "test_step_list": [{
                                                                                                "test_step_number": f"step{str(relationship_properties_level_2_sorting)[1:-1]}",
                                                                                                "test_step_content": str(
                                                                                                    test_step)[1:-1],
                                                                                                "test_step_parameter": str(
                                                                                                    relationship_properties_parameter)[
                                                                                                                       1:-1]}]})
                if first_tag == False:
                    result_list.append({"method_name": str(method_name)[1:-1], "test_step_sub_method_list": [
                        {"sub_method_name": str(sub_method_name)[1:-1], "test_step_list": [
                            {"test_step_number": f"step{str(relationship_properties_level_2_sorting)[1:-1]}",
                             'test_step_content': str(test_step)[1:-1],
                             "test_step_parameter": str(relationship_properties_parameter)[1:-1]}]}],"check_item_sub_method_list":[]})

        with connection_pool.session_context('root', 'nebula') as session:
            session.execute('USE digitalSE_V5')
            query = f'MATCH (p:PN)-[e1]->(d:PRV_DOCUMENT)-[e2]-(i:TEST_CHECK_ITEMS) WHERE id(p) == "{common_pn_result[0]}" RETURN e2,i,toInteger(e2.`level_1_sorting `) AS l1, toInteger(e2.`level_2_sorting`) AS l2 ORDER BY l1,l2;'
            result = session.execute(query)
            for row in result:
                check = row.values()[1]
                if check:
                    check_obj = check.as_node()
                    check_properties = check_obj.properties()
                    check_method_name = check_properties['parent_test_method']
                    check_sub_method_name = check_properties['subparent_test_method']
                    check_item = check_properties['name']
                #                     print(check_method_name,check_sub_method_name,check_item)
                relationship = row.values()[0]
                if relationship:
                    relationship_obj = relationship.as_relationship()
                    check_relationship_properties = relationship_obj.properties()
                    check_relationship_properties_min = check_relationship_properties["min"]
                    check_relationship_properties_nomial = check_relationship_properties["nomial"]
                    check_relationship_properties_max = check_relationship_properties["max"]

                    for first_result_dict in result_list:
                        if first_result_dict["method_name"].strip() == str(check_method_name)[1:-1].strip():
                            check_second_tag = False
                            for second_dict in first_result_dict["check_item_sub_method_list"]:
                                if second_dict["sub_method_name"] == str(check_sub_method_name)[1:-1]:
                                    second_dict["check_item_list"].append(
                                        {"check_item_name": str(check_item)[1:-1],
                                         "min": str(check_relationship_properties_min)[1:-1],
                                         "nomial": str(check_relationship_properties_nomial)[1:-1],
                                         "max": str(check_relationship_properties_max)[1:-1]})
                                    check_second_tag = True
                                    break
                            if check_second_tag == False:
                                first_result_dict["check_item_sub_method_list"].append(
                                    {"sub_method_name": str(check_sub_method_name)[1:-1], "check_item_list": [
                                        {"check_item_name": str(check_item)[1:-1],
                                         "min": str(check_relationship_properties_min)[1:-1],
                                         "nomial": str(check_relationship_properties_nomial)[1:-1],
                                         "max": str(check_relationship_properties_max)[1:-1]}]})
            #

    return result_list


def extract_query_instruction(user_question):
    # 加载数据
    # end_data_df = pd.read_excel('./csv3/comb_add_step.xlsx')
    # 获取当前文件的绝对路径
    current_file_path = os.path.abspath(__file__)
    # 获取当前文件的根目录
    root_directory = os.path.dirname(current_file_path)
    end_data_df = pd.read_excel(os.path.join(root_directory, 'comb_add_step.xlsx'))
    column_name = ', '.join(end_data_df.columns)
    extraction_instruction = """
你将接收一个与表格数据相关的问题，请你从中提取出标准化的“查询指令”，用于程序自动识别和过滤数据。

⚠️ 输出格式必须是严格的 JSON（不要添加自然语言说明），例如：

{
  "query_conditions": [
    {"field": "字段1", "operator": "=", "value": "值1"},
    {"field": "字段2", "operator": "包含", "value": "值2"}
  ],
  "return_fields": ["字段3", "字段4"],
  "distinct": true,
  "sort_by": {"field": "xxx", "order": "升序"},
  "limit": 10
}

同时所有字段必须和实际表格数据的列名%s匹配，避免到表格中找不到的情况，
表格字段说明如下：
    pn_vid：产品虚拟ID（例如“pn1”）
    module_vid：组成零部件的模块虚拟ID
    module_details：组成零部件的模块具体名称或型号（例如“MCU_C600_12V”）
    module_type：组成零部件的模块类型（例如“MCU”、“DCDC”、“PDU”）
    pn_number：实际零件编号（例如“0437CX001F”）
    pn_name：零件名称（例如“低压液剂”）
    doc_vid：文档虚拟ID
    series_vid：产品系虚拟ID
    series_name：产品系名称（例如“Xin1-M”、“Xin_S”）
    product_vid：产品虚拟ID
    product_name：产品名称（例如“Xin1”）
    doc_name：文档名称（例如“0437P001FN-PRV_…”）
    doc_time：文档创建时间（例如“20241209”）
    customer_vid：客户虚拟ID
    customer_name：客户名称（例如“JMC E820”）
    reference_parameter_for_step:测试步骤中的相关参数
    level1_sort_number_for_step:测试步骤的第一层排序
    level2_sort_number_for_step:测试步骤在第一层排序下的第二层排序
    step_vid:测试步骤虚拟ID
    step_level1:测试步骤所属的大标题
    step_level2:测试步骤所属的子标题
    step_level3:测试步骤的内容


下面是用户的问题，请提取指令：
""" % column_name

    messages = [
        {"role": "system", "content": extraction_instruction},
        {"role": "user", "content": user_question}
    ]

    try:
        response = azure_client.chat.completions.create(
            model="gpt-4o",
            messages=messages,
            temperature=0.1,
            stream=False
        )
        content = response.choices[0].message.content.strip()

        # 防止代码块包裹
        if content.startswith("```json"):
            content = content.replace("```json", "").replace("```", "").strip()

        return "ok", json.loads(content)

    except json.JSONDecodeError:
        print("原始内容：", content)
        return "nok", "返回内容不是有效的 JSON 格式"
        # raise ValueError("返回内容不是有效的 JSON 格式")
    except Exception as e:
        print(f"调用模型失败: {e}")
        return "nok", f"调用模型失败,报错详情如下: {e}"
        # raise RuntimeError(f"调用模型失败: {e}")


def llm_chat(task_id, user_question):
    # # 加载数据
    # end_data_df = pd.read_excel('./csv3/comb_add_step.xlsx',dtype=str)
    # 获取当前文件的绝对路径
    current_file_path = os.path.abspath(__file__)
    # 获取当前文件的根目录
    root_directory = os.path.dirname(current_file_path)
    end_data_df = pd.read_excel(os.path.join(root_directory, 'comb_add_step.xlsx'), dtype=str)
    end_data_df.fillna('', inplace=True)
    # 确保用户输入是字符串类型
    if not isinstance(user_question, str):
        user_question = str(user_question)  # 如果不是字符串，强制转换为字符串

    # 用 LLM 结构化提取查询
    try:
        status, structured_query = extract_query_instruction(user_question)
        if status == "ok":
            # === 执行查询 ===
            df = end_data_df.copy()

            # 处理条件
            for cond in structured_query.get("query_conditions", []):
                field = cond["field"]
                val = cond["value"]
                if cond["operator"] == "=":
                    df = df[df[field] == val]
                elif cond["operator"] == "包含":
                    df = df[df[field].astype(str).str.contains(val, na=False)]

            # 去重
            if structured_query.get("distinct"):
                df = df.drop_duplicates(subset=structured_query["return_fields"])

            # 排序
            sort = structured_query.get("sort_by")
            if sort:
                sort_field = sort["field"]
                df = df.sort_values(by=sort_field, ascending=(sort["order"] == "升序"))

            # 限制
            if structured_query.get("limit"):
                df = df.head(structured_query["limit"])

            # 返回字段
            # return_fields = structured_query["return_fields"]
            # for cond in structured_query.get("query_conditions", []):
            #     field = cond["field"]
            #     return_fields.append(field)
            return_fields = structured_query.get("return_fields", [])
            # 添加查询字段，但避免重复
            query_fields = [cond["field"] for cond in structured_query.get("query_conditions", [])]
            return_fields = list(set(return_fields + query_fields))

            final_df = df[return_fields].reset_index(drop=True)
            result_dict = {'columns': list(final_df.columns), 'rows': []}
            for i in range(len(final_df)):
                row_data = dict(final_df.loc[i])
                result_dict['rows'].append(row_data)

            table_message = {'connectionID': task_id,
                             'category': 'table', 'from': '', 'to': '',
                             'message': json.dumps(result_dict, ensure_ascii=False),
                             'remarks': json.dumps({'paragraph_start': 1,
                                                    'response_end': 0})}
            try:
                asyncio.run(websocket_client(table_message))
            except:
                time.sleep(0.5)
                asyncio.run(websocket_client(table_message))

            # Step 2: 生成 summary 的 prompt
            # 获取流式输出
            paragraph_start = 1
            response_text = ''
            summary_prompt = f"""
            你是一名测试数据分析助手。请根据以下数据内容，用简洁中文总结核心信息，不要逐行复述：
            字段：{result_dict['columns']}
            内容：
            {json.dumps(result_dict['rows'], ensure_ascii=False, indent=2)}
            """
            chat_prompt = [
                {
                    "role": "system",
                    "content": [
                        {
                            "type": "text",
                            "text": summary_prompt
                        }
                    ]
                }
            ]
            # Include speech result if speech is enabled
            conversation = chat_prompt
            response = azure_client.chat.completions.create(
                model="gpt-4o",  # 使用 GPT-4 模型
                messages=conversation,
                temperature=0,  # 设置生成的温度
                stream=True  # 启用流式输出
            )
            for chunk in response:
                if chunk.choices:
                    text = chunk.choices[0].delta.content
                    if text is not None:
                        text_message = {'connectionID': task_id,
                                        'category': 'text', 'from': '', 'to': '',
                                        'message': text,
                                        'remarks': json.dumps({'paragraph_start': paragraph_start,
                                                               'response_end': 0})}
                        try:
                            asyncio.run(websocket_client(text_message))
                        except:
                            time.sleep(0.5)
                            asyncio.run(websocket_client(text_message))

                        paragraph_start = 0
                        response_text += text

            text_message = {'connectionID': task_id,
                            'category': 'text', 'from': '', 'to': '',
                            'message': ' ',
                            'remarks': json.dumps({'paragraph_start': 0,
                                                   'response_end': 1})}

            asyncio.run(websocket_client(text_message))

            # return result_dict
    #         final_df.to_excel("result_output.xlsx", index=False)
    #         print("✅ 结果已保存到 result_output.xlsx")

        else:
            error_message = {'connectionID': task_id,
                             'category': 'error', 'from': '', 'to': '',
                             'message': json.dumps(structured_query, ensure_ascii=False),
                             'remarks': json.dumps({'paragraph_start': 1,
                                                    'response_end': 1})}
            try:
                asyncio.run(websocket_client(error_message))
            except:
                time.sleep(0.5)
                asyncio.run(websocket_client(error_message))
    except Exception as e:
        print(f"⚠️ 查询或处理失败：{e}")
        error_message = {'connectionID': task_id,
                         'category': 'error', 'from': '', 'to': '',
                         'message': json.dumps(f"⚠️ 查询或处理失败：{e}", ensure_ascii=False),
                         'remarks': json.dumps({'paragraph_start': 1,
                                                'response_end': 1})}

        try:
            asyncio.run(websocket_client(error_message))
        except:
            time.sleep(0.5)
            asyncio.run(websocket_client(error_message))
        # return {}  # 返回空表以防出错


###################################################################################################

                             ##       WUJ AUTO PRV       ##

###################################################################################################

# def auto_prv(task_id,pdf_path):
#     def clean_llm_json(json_str):
#         # 去掉前后的 markdown 代码块标识符 ```json ... ```
#         if json_str.startswith("```"):
#             json_str = re.sub(r"^```(?:json)?\n", "", json_str)
#             json_str = re.sub(r"\n```$", "", json_str)
#         return json_str.strip()
#
#     # 图像提取
#     def convert_pdf_to_images(pdf_path):
#         doc = fitz.open(pdf_path)
#         image_paths = []
#         for i in range(len(doc)):
#             page = doc[i]
#             pix = page.get_pixmap(dpi=200)
#             image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{i + 1}.png"
#             #github存储 image
#             # 获取当前文件的绝对路径
#             current_file_path = os.path.abspath(__file__)
#             # 获取当前文件的根目录
#             root_directory = os.path.dirname(current_file_path)
#             local_img_path = os.path.join(root_directory, 'prv_pdf_images', f"{image_filename}")
#             print("local_img_path: ", local_img_path)
#             pix.save(local_img_path)
#             # with open(local_img_path, 'wb') as file:
#             #     file.write(pix)
#             # 文件服务器存储image
#             with open(local_img_path, "rb") as image:
#                 files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
#                 upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
#                 response = requests.post(upload_url, files=files, verify=False)
#                 image_path = response.json()['data'][0]['url']
#
#             image_paths.append(image_path)
#         return image_paths
#
#     # GPT-4o 调用
#     def extract_data_from_image(image_path):
#         response = requests.get(image_path, verify=False)
#         if response.status_code == 200:
#             image_bytes = response.content
#             base64_img = base64.b64encode(image_bytes).decode('utf-8')
#         else:
#             raise Exception(f"Failed to fetch image. Status code: {response.status_code}")
#         # with open(image_path, "rb") as f:
#         #     image_bytes = f.read()
#         #
#         # base64_img = base64.b64encode(image_bytes).decode('utf-8')
#         # 提取字段 Prompt 模板
#         PROMPT_TEMPLATE = """你将看到一张图片，图片中可能包含 EOL 测试步骤。如果图片上有以 "TAS" 开头并以五位数字结尾的编码（例如 TAS20678），
#         以及包含如下字段的无框表格：
#
#         - Action
#         - LSL
#         - USL
#         - Unit
#         - Circuit
#         - Parameter
#         - Description
#
#         请你提取该图片中的所有测试步骤，每个步骤提取成一个字典，字段如下：
#
#         {
#           "TAS ID": "...",
#           "Action": "...",
#           "LSL": "...",
#           "USL": "...",
#           "Unit": "...",
#           "Circuit": "...",
#           "Parameter": "...",
#           "Description": "..."
#         }
#
#         并将这些字典放入一个列表中返回，最终输出为标准 JSON 格式：
#
#         [
#           {...},
#           {...},
#           ...
#         ]
#
#         ⚠️ 注意事项：
#         - 一张图片中可能包含多个测试步骤，务必提取**全部**。
#         - 字段必须按照上面顺序，不可缺失。
#         - 若某些字段为空，请也保留字段并赋值为空字符串 ""。
#         - 返回结果必须是合法 JSON，**不要添加任何解释说明文字或多余注释**。
#         """
#         response = azure_client.chat.completions.create(
#             model="gpt-4o",
#             messages=[
#                 {
#                     "role": "system",
#                     "content": "你是一个图像信息提取专家，擅长从截图中读取结构化信息。"
#                 },
#                 {
#                     "role": "user",
#                     "content": [
#                         {"type": "text", "text": PROMPT_TEMPLATE},
#                         {
#                             "type": "image_url",
#                             "image_url": {
#                                 "url": f"data:image/png;base64,{base64_img}"
#                             }
#                         }
#                     ]
#                 }
#             ],
#             temperature=0
#         )
#         return response.choices[0].message.content
#
#     image_paths = convert_pdf_to_images(pdf_path)
#     records = []
#     for img_path in tqdm(image_paths, desc=f"→ Extracting from {pdf_path}", leave=False):
#
#         try:
#             json_str = extract_data_from_image(img_path)
#             print("🔍 LLM 原始返回内容：", repr(json_str))
#
#             json_str_clean = clean_llm_json(json_str)
#             print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容
#
#             if json_str_clean.strip():  # 防止空字符串
#                 data_list = json.loads(json_str_clean)
#                 # 异步传输每个image和content
#                 result_dict = {'page_image_path': img_path, 'page_image_content': data_list}
#                 prv_message = {'connectionID': task_id,
#                                  'category': 'prv_image', 'from': '', 'to': '',
#                                  'message': json.dumps(result_dict, ensure_ascii=False),
#                                  'remarks': json.dumps({'paragraph_start': 1,
#                                                         'response_end': 0})}
#                 try:
#                     asyncio.run(websocket_client(prv_message))
#                 except:
#                     time.sleep(0.5)
#                     asyncio.run(websocket_client(prv_message))
#
#                 # 不为空再记录 后续存到excel
#                 if data_list:
#                     records.extend(data_list)
#             else:
#                 print(f"⚠️ 空内容跳过解析：{img_path}")
#
#         except Exception as e:
#             print(f"❌ Error parsing {img_path}: {e}")
#             # print(f"⚠️ 原始返回内容为： {repr(json_str)}")
#
#     # 异步传输结束信号空字符串
#     prv_message = {'connectionID': task_id,
#                      'category': 'text', 'from': '', 'to': '',
#                      'message': '',
#                      'remarks': json.dumps({'paragraph_start': 0,
#                                             'response_end': 1})}
#     try:
#         asyncio.run(websocket_client(prv_message))
#     except:
#         time.sleep(0.5)
#         asyncio.run(websocket_client(prv_message))
#
#     # 结果生成excel存储到github
#     if records:
#         df = pd.DataFrame(records)
#         excel_name = os.path.basename(pdf_path).replace('.pdf', '.xlsx')
#         # 获取当前文件的绝对路径
#         current_file_path = os.path.abspath(__file__)
#         # 获取当前文件的根目录
#         root_directory = os.path.dirname(current_file_path)
#         df.to_excel(os.path.join(root_directory, 'prv_pdf_images', f"{excel_name}"), index=False)
#     # return records
#
#
#
# def auto_prv_improve(task_id,pdf_path):
#     # original_pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
#     # print("save original_pdf_name", original_pdf_name)
#     download_excel_path_dict = {}
#     def clean_llm_json(json_str):
#         # 去掉前后的 markdown 代码块标识符 ```json ... ```
#         if json_str.startswith("```"):
#             json_str = re.sub(r"^```(?:json)?\n", "", json_str)
#             json_str = re.sub(r"\n```$", "", json_str)
#         return json_str.strip()
#
#         # GPT-4o 调用
#
#     def extract_data_from_image(image_path):
#         response = requests.get(image_path, verify=False)
#         if response.status_code == 200:
#             image_bytes = response.content
#             base64_img = base64.b64encode(image_bytes).decode('utf-8')
#         else:
#             raise Exception(f"Failed to fetch image. Status code: {response.status_code}")
#         # with open(image_path, "rb") as f:
#         #     image_bytes = f.read()
#         #
#         # base64_img = base64.b64encode(image_bytes).decode('utf-8')
#         # 提取字段 Prompt 模板
#         PROMPT_TEMPLATE = """你将看到一张图片，图片中可能包含 EOL 测试步骤。如果图片上有以 "TAS" 开头并以五位数字结尾的编码（例如 TAS20678），以及编号旁的灰色字体为Station，
#             以及编号和灰色字体下的为Program，以及包含如下字段的无框表格：
#
#                - Action
#                - LSL
#                - USL
#                - Unit
#                - Circuit
#                - Parameter
#                - Description
#
#                请你提取该图片中的所有测试步骤，每个步骤提取成一个字典，字段如下：
#
#                {
#                  "TAS ID": "...",
#                  "Station": "...",
#                  "Program": "...",
#                  "Action": "...",
#                  "LSL": "...",
#                  "USL": "...",
#                  "Unit": "...",
#                  "Circuit": "...",
#                  "Parameter": "...",
#                  "Description": "..."
#                }
#
#                并将这些字典放入一个列表中返回，最终输出为标准 JSON 格式：
#
#                [
#                  {...},
#                  {...},
#                  ...
#                ]
#
#                ⚠️ 注意事项：
#                - 如果图片内容不符合上面描述的格式，请直接返回空。
#                - 一张图片中可能包含多个测试步骤，务必提取**全部**。
#                - 字段必须按照上面顺序，不可缺失。
#                - 若某些字段为空，请也保留字段并赋值为空字符串 ""，不要串行。
#                - 返回结果必须是合法 JSON，**不要添加任何解释说明文字或多余注释**。
#                """
#
#         response = azure_client.chat.completions.create(
#             model="gpt-4o",
#             messages=[
#                 {
#                     "role": "system",
#                     "content": "你是一个图像信息提取专家，擅长从截图中读取结构化信息。"
#                 },
#                 {
#                     "role": "user",
#                     "content": [
#                         {"type": "text", "text": PROMPT_TEMPLATE},
#                         {
#                             "type": "image_url",
#                             "image_url": {
#                                 "url": f"data:image/png;base64,{base64_img}"
#                             }
#                         }
#                     ]
#                 }
#             ],
#             temperature=0
#         )
#         return response.choices[0].message.content
#
#         # 获取当前文件的绝对路径
#
#     current_file_path = os.path.abspath(__file__)
#     # 获取当前文件的根目录
#     root_directory = os.path.dirname(current_file_path)
#     records = []
#     doc = fitz.open(pdf_path)
#     for i in range(len(doc)):
#         page = doc[i]
#         pix = page.get_pixmap(dpi=200)
#         image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{i + 1}.png"
#         # docker 存储 image
#         local_img_path = os.path.join(root_directory, 'prv_pdf_images', f"{image_filename}")
#         print("local_img_path: ", local_img_path)
#         pix.save(local_img_path)
#         # with open(local_img_path, 'wb') as file:
#         #     file.write(pix)
#         # 文件服务器存储image
#         with open(local_img_path, "rb") as image:
#             files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
#             upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
#             response = requests.post(upload_url, files=files, verify=False)
#             image_path = response.json()['data'][0]['url']
#
#             try:
#                 json_str = extract_data_from_image(image_path)
#                 print("🔍 LLM 原始返回内容：", repr(json_str))
#
#                 json_str_clean = clean_llm_json(json_str)
#                 print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容
#
#                 if json_str_clean.strip():  # 防止空字符串
#                     data_list = json.loads(json_str_clean)
#                     print("返回的prv表格: ", data_list)
#
#                     page_testing_matrix = []
#
#                     # 收集本页所有的 tas id
#                     tas_id_list = []
#                     for data in data_list:
#                         tas_id_list.append(data['TAS ID'])
#                     tas_id_unique_list = list(set(tas_id_list))
#
#                     # 一个 tas id 下会有一行或多行 需要by tas id分组使用
#                     for tas_id in tas_id_unique_list:
#                         in_use_data_list = []
#                         for data_list_row in data_list:
#                             if data_list_row['TAS ID'] == tas_id:
#                                 in_use_data_list.append(data_list_row)
#                         print('TES ID: ', tas_id, 'Station: ', in_use_data_list[0]['Station'], 'RELATED DATA LIST: ', in_use_data_list)
#                         testing_matrix_dict = {}
#                         data = in_use_data_list[0]
#                         # 场景一 ：【SW_Vers】
#                         if '[SW_Vers]' in data['Program']:
#                             sw_ver_testing_matrix = [
#                                 {"Test Id": ".01", "Input Message": "sUTIL_GetVersionData", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "0",
#                                  "HiLimit": "0", "Unit": "Dec", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rg_Value", "speParamString": "VERSION_KEY",
#                                  "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".02", "Input Message": "sUTIL_VerifyResult", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rg_Value", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".03", "Input Message": "sDoIP_SND_RCV_RC", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "10",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rg_Value", "speParamString": "", "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".04", "Input Message": "sUTIL_VerifyResult", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rm_DoIP_Param1", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": "", "Input Message": "sUTIL_Compare_String", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "",
#                                  "speParamString": "STRING1 {@Rg_Value@}\nSTRING2 {@Rm_DoIP_Param1@}\nCOMPARE_STRING {String1_String2} TO_UPPER {0}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                             print("***场景一匹配：SW_Vers***")
#                             try:
#                                 # step1 根据prv的tasid填写所有test id
#                                 tas_id = data['TAS ID']
#                                 tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                 for index, matrix_row in enumerate(sw_ver_testing_matrix):
#                                     if index == 4:
#                                         matrix_row['Test Id'] = tas_id
#                                     else:
#                                         matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#                                 # step2 根据prv的paramter中的TX填写第三行的speSend
#                                 parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 sw_ver_testing_matrix[2]['speSend'] = '"' + parameter_tx + '"'
#                                 # step3 根据prv的paramter中的RX填写第三行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ', '').replace('zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 sw_ver_testing_matrix[2]['LoLimit'] = parameter_rx
#                                 sw_ver_testing_matrix[2]['HiLimit'] = parameter_rx
#                                 sw_ver_testing_matrix[2]['speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#                                 # step4 根据prv的LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                 lsl = data['LSL']
#                                 usl = data['USL']
#                                 unit = data['Unit']
#                                 sw_ver_testing_matrix[4]['LoLimit'] = lsl
#                                 sw_ver_testing_matrix[4]['HiLimit'] = usl
#                                 sw_ver_testing_matrix[4]['Unit'] = unit
#                                 # step5 根据prv中的program name 填写第一行的speParamString
#                                 program_name = data['Program'].replace('[SW_Vers]', '').strip()
#                                 sw_ver_testing_matrix[0]['speParamString'] = sw_ver_testing_matrix[0][
#                                                                                   'speParamString'] + "{ASS_" + program_name + "}"
#                                 # step6 根据prv中的program name 填写每行Description
#                                 sw_ver_testing_matrix[0]['Description'] = "Read " + program_name + " from dataface"
#                                 sw_ver_testing_matrix[1]['Description'] = "Parse " + program_name + " to String"
#                                 sw_ver_testing_matrix[2]['Description'] = "Read " + program_name + " from DUT"
#                                 sw_ver_testing_matrix[3]['Description'] = "Parse " + program_name + " from DUT"
#                                 sw_ver_testing_matrix[4]['Description'] = "Check " + program_name + " with dataface"
#                                 # 修改完成
#                                 print("sw_ver_testing_matrix： ", sw_ver_testing_matrix)
#                                 # 存入字典
#                                 testing_matrix_dict['station'] = data['Station']
#                                 testing_matrix_dict['category'] = program_name
#                                 testing_matrix_dict['testing_content'] = sw_ver_testing_matrix
#                                 # 加入list
#                                 page_testing_matrix.append(testing_matrix_dict)
#                             except Exception as e:
#                                 print(f"❌ Error transfer prv content to testing matrix in sw_vers category: {e}")
#
#                         # 场景二 ：【GPIO PGOOD】
#                         if '[GPIO PGOOD]' in data['Program']:
#                             gpio_pgood_testing_matrix = [
#                                 {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC", "Description": "Read DUT PGOOD Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "10",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".02", "Input Message": "sUTIL_VerifyResult", "Description": "Parse DUT PGOOD Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rm_DoIP_Param1", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": "", "Input Message": "sUTIL_ParseTASResultAsHex", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                             print("***场景二匹配：GPIO PGOOD***")
#                             try:
#                                 # step1 根据prv的tasid填写所有test id
#                                 tas_id = data['TAS ID']
#                                 tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                 for index, matrix_row in enumerate(gpio_pgood_testing_matrix):
#                                     if index == 2:
#                                         matrix_row['Test Id'] = tas_id
#                                     else:
#                                         matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#                                 # step2 根据prv的paramter中的TX填写第一行的speSend
#                                 parameter_tx = data['Parameter'].split('RX')[0].replace('TX:', '').replace('|', '').strip()
#                                 # parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 gpio_pgood_testing_matrix[0]['speSend'] = '"' + parameter_tx + '"'
#                                 # step3 根据prv的paramter中的RX填写第一行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('zz')[0].split('RX:')[1].strip()
#                                 # parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                 #                                                                           '').replace(
#                                 #     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 gpio_pgood_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 gpio_pgood_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 gpio_pgood_testing_matrix[0][
#                                     'speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#                                 # step4 根据prv的LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                 lsl = data['LSL']
#                                 usl = data['USL']
#                                 unit = data['Unit']
#                                 gpio_pgood_testing_matrix[2]['LoLimit'] = lsl
#                                 gpio_pgood_testing_matrix[2]['HiLimit'] = usl
#                                 gpio_pgood_testing_matrix[2]['Unit'] = unit
#                                 # step5 根据prv中的description 填写最后一行的speReference
#                                 match = re.search(r'check\s*byte\s*(\d+)', data['Description'])
#                                 if match:
#                                     number = int(match.group(1))
#                                     print(number)
#                                     final_number = number * 3
#                                     gpio_pgood_testing_matrix[2]['speReference'] = '"%s*' + str(
#                                         final_number) + '%s2"'
#
#                                 # step6 将第二行的test id 填写入最后一行的speResult
#                                 gpio_pgood_testing_matrix[2]['speResult'] = gpio_pgood_testing_matrix[1]['Test Id']
#
#                                 # step7 根据prv中的program name 填写最后一行Description
#                                 program_name = data['Program'].replace('[GPIO PGOOD]', '').strip()
#                                 gpio_pgood_testing_matrix[2]['Description'] = "Check DUT " + program_name + " status"
#
#                                 print("gpio_pgood_testing_matrix： ", gpio_pgood_testing_matrix)
#                                 # 存入字典
#                                 testing_matrix_dict['station'] = data['Station']
#                                 testing_matrix_dict['category'] = program_name
#                                 testing_matrix_dict['testing_content'] = gpio_pgood_testing_matrix
#                                 print("check one part testing_matrix_dict: ", testing_matrix_dict)
#                                 # 加入list
#                                 page_testing_matrix.append(testing_matrix_dict)
#                             except Exception as e:
#                                 print(f"❌ Error transfer prv content to testing matrix in GPIO_PGOOD category: {e}")
#
#                         # 场景三 ：【Thermal】
#                         if '[Thermal]' in data['Program']:
#
#                             thermal_testing_matrix = [
#                                 {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC", "Description": "Read DUT Thermal Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "10",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".02", "Input Message": "sUTIL_VerifyResult", "Description": "Parse DUT Thermal Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rm_DoIP_Param1", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": "", "Input Message": "sUTIL_ParseTASResultAsHex", "Description": "",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                             print("***场景三匹配：Thermal***")
#                             try:
#                                 # step1 根据prv的tasid填写所有test id
#                                 tas_id = data['TAS ID']
#                                 tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                 for index, matrix_row in enumerate(thermal_testing_matrix):
#                                     if index == 2:
#                                         matrix_row['Test Id'] = tas_id
#                                     else:
#                                         matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#                                 # step2 根据prv的paramter中的TX填写第一行的speSend
#                                 parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 thermal_testing_matrix[0]['speSend'] = '"' + parameter_tx + '"'
#                                 # step3 根据prv的paramter中的RX填写第一行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 thermal_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 thermal_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 thermal_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#                                 # step4 根据prv的LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                 lsl = data['LSL']
#                                 usl = data['USL']
#                                 unit = data['Unit']
#                                 thermal_testing_matrix[2]['LoLimit'] = lsl
#                                 thermal_testing_matrix[2]['HiLimit'] = usl
#                                 thermal_testing_matrix[2]['Unit'] = unit
#                                 # step5 根据prv中的description 填写最后一行的speReference
#                                 match = re.search(r'check\s*byte\s*(\d+)\s*[-~]\s*(\d+)', data['Description'],
#                                                   re.IGNORECASE)
#                                 if match:
#                                     byte_start = int(match.group(1))
#                                     byte_end = int(match.group(2))
#                                     print(
#                                         f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#                                     first_number = byte_start * 3
#                                     byte_length = int(byte_end - byte_start + 1)
#                                     second_number = int(byte_length * 3 - 1)
#                                     thermal_testing_matrix[2]['speReference'] = '"%s*' + str(
#                                         first_number) + '%s' + str(second_number) + '"'
#                                 else:
#                                     print("No match found.")
#
#                                 # step6 将第二行的test id 填写入最后一行的speResult
#                                 thermal_testing_matrix[2]['speResult'] = thermal_testing_matrix[1]['Test Id']
#
#                                 # step7 根据prv中的program name 填写最后一行Description
#                                 program_name = data['Program'].replace('[Thermal]', '').strip()
#                                 thermal_testing_matrix[2]['Description'] = "Check DUT " + program_name + " status"
#
#                                 print("thermal_testing_matrix： ", thermal_testing_matrix)
#                                 # 存入字典
#                                 testing_matrix_dict['station'] = data['Station']
#                                 testing_matrix_dict['category'] = program_name
#                                 testing_matrix_dict['testing_content'] = thermal_testing_matrix
#                                 # 加入list
#                                 page_testing_matrix.append(testing_matrix_dict)
#                             except Exception as e:
#                                 print(f"❌ Error transfer prv content to testing matrix in Thermal category: {e}")
#
#                         # 场景四：【FAN】 Fan on
#                         if '[FAN]' in data['Program'] and 'FAN ON' in data['Program']:
#                             data2 = in_use_data_list[1]
#                             fan_on_testing_matrix = [
#                                 {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC", "Description": "Set FAN ON test",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "10",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".02", "Input Message": "sUTIL_Delay", "Description": "Waiting 1s for DUT FAN ON test",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "0",
#                                  "HiLimit": "0", "Unit": "Dec", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "1",
#                                  "speTimeOut": "", "speDelay": "1000", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "",
#                                  "speParamString": "DELAY_DESCRIPTION {Waiting DUT Ready}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".03", "Input Message": "sDoIP_SND_RCV_RC", "Description": "Read DUT FAN ON test ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".04", "Input Message": "sUTIL_VerifyResult", "Description": "Parse DUT FAN ON test ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rm_DoIP_Param1", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": "", "Input Message": "sUTIL_ParseTASResultAsHex", "Description": "Check DUT FAN ON test ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                             print("***场景四匹配：FAN ON ***")
#                             try:
#                                 # step1 根据prv的tasid填写所有test id
#                                 tas_id = data['TAS ID']
#                                 tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                 for index, matrix_row in enumerate(fan_on_testing_matrix):
#                                     if index == 4:
#                                         matrix_row['Test Id'] = tas_id
#                                     else:
#                                         matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#
#                                 # step2 根据prv的paramter中第一行的TX填写第一行的speSend 并且需要替换xx
#                                 parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 match_replace_xx_value = re.search(r'set\s*byte0:\s*0x([0-9a-fA-F]+)',
#                                                                    data['Description'], re.IGNORECASE)
#                                 if match_replace_xx_value:
#                                     replace_xx_value = str(match_replace_xx_value.group(1))
#                                     fan_on_testing_matrix[0]['speSend'] = '"' + parameter_tx.replace('xx',
#                                                                                                      replace_xx_value) + '"'
#                                 else:
#                                     print("No match_replace_xx_value found in FAN ON.")
#
#                                 # step2.1 根据prv的paramter中第二行的TX填写第三行的speSend
#                                 parameter_tx = data2['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 fan_on_testing_matrix[2]['speSend'] = '"' + parameter_tx + '"'
#
#                                 # step3 根据prv的paramter中第二行的RX填写第一行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 fan_on_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 fan_on_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 fan_on_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + '"'
#
#                                 # step3.1 根据prv的paramter中第二行的RX填写第三行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 fan_on_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 fan_on_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 fan_on_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#
#                                 # step4 根据prv的第二行LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                 lsl = data2['LSL']
#                                 usl = data2['USL']
#                                 unit = data2['Unit']
#                                 fan_on_testing_matrix[4]['LoLimit'] = lsl
#                                 fan_on_testing_matrix[4]['HiLimit'] = usl
#                                 fan_on_testing_matrix[4]['Unit'] = unit
#                                 # step5 根据prv中第二行的description 填写最后一行的speReference
#                                 match = re.search(r'check\s*byte\s*(\d+)\s*[-~]\s*(\d+)', data2['Description'],
#                                                   re.IGNORECASE)
#                                 if match:
#                                     byte_start = int(match.group(1))
#                                     byte_end = int(match.group(2))
#                                     print(
#                                         f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#                                     if byte_start == 0:
#                                         # check从头开始不需要标记位置 直接填写需要check的长度 byte_end 就是byte length
#                                         byte_length = byte_end
#                                         number = int(byte_length * 3 - 1)
#                                         fan_on_testing_matrix[4]['speReference'] = '"%s' + str(number) + '"'
#
#                                     else:
#                                         # first_number代表开始check的位置 second_number代表check的长度
#                                         first_number = byte_start * 3
#                                         byte_length = int(byte_end - byte_start + 1)
#                                         second_number = int(byte_length * 3 - 1)
#                                         fan_on_testing_matrix[4]['speReference'] = '"%s*' + str(
#                                             first_number) + '%s' + str(second_number) + '"'
#                                 else:
#                                     print("No match found.")
#
#                                 # step6 将第二行的test id 填写入最后一行的speResult
#                                 fan_on_testing_matrix[4]['speResult'] = fan_on_testing_matrix[3]['Test Id']
#
#                                 #             # step7 根据prv中的program name 填写最后一行Description
#                                 #             program_name = data['Program'].replace('[Thermal]','').strip()
#                                 #             thermal_testing_matrix[2]['Description'] = "Check DUT " + program_name + " status"
#
#                                 print("fan_on_testing_matrix： ", fan_on_testing_matrix)
#                                 # 存入字典
#                                 testing_matrix_dict['station'] = data['Station']
#                                 testing_matrix_dict['category'] = 'FAN ON test'
#                                 testing_matrix_dict['testing_content'] = fan_on_testing_matrix
#                                 # 加入list
#                                 page_testing_matrix.append(testing_matrix_dict)
#                             except Exception as e:
#                                 print(f"❌ Error transfer prv content to testing matrix in fan on category: {e}")
#
#                         # 场景五：【FAN】 Fan OFF
#                         if '[FAN]' in data['Program'] and 'FAN OFF' in data['Program']:
#                             data2 = in_use_data_list[1]
#                             fan_off_testing_matrix = [
#                                 {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC",
#                                  "Description": "Set FAN OFF test",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "10",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".02", "Input Message": "sUTIL_Delay",
#                                  "Description": "Waiting 1s for DUT FAN OFF test",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "0",
#                                  "HiLimit": "0", "Unit": "Dec", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "1",
#                                  "speTimeOut": "", "speDelay": "1000", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "DELAY_DESCRIPTION {Waiting DUT Ready}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".03", "Input Message": "sDoIP_SND_RCV_RC",
#                                  "Description": "Read DUT FAN OFF test ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".04", "Input Message": "sUTIL_VerifyResult",
#                                  "Description": "Parse DUT FAN OFF test ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rm_DoIP_Param1", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": "", "Input Message": "sUTIL_ParseTASResultAsHex",
#                                  "Description": "Check DUT FAN OFF test ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                             print("***场景五匹配：FAN OFF ***")
#                             try:
#                                 # step1 根据prv的tasid填写所有test id
#                                 tas_id = data['TAS ID']
#                                 tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                 for index, matrix_row in enumerate(fan_off_testing_matrix):
#                                     if index == 4:
#                                         matrix_row['Test Id'] = tas_id
#                                     else:
#                                         matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#
#                                 # step2 根据prv的paramter中第一行的TX填写第一行的speSend 并且需要替换xx
#                                 parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 match_replace_xx_value = re.search(r'set\s*byte0:\s*0x([0-9a-fA-F]+)',
#                                                                    data['Description'], re.IGNORECASE)
#                                 if match_replace_xx_value:
#                                     replace_xx_value = str(match_replace_xx_value.group(1))
#                                     fan_off_testing_matrix[0]['speSend'] = '"' + parameter_tx.replace('xx',
#                                                                                                       replace_xx_value) + '"'
#                                 else:
#                                     print("No match_replace_xx_value found in FAN OFF.")
#
#                                 # step2.1 根据prv的paramter中第二行的TX填写第三行的speSend
#                                 parameter_tx = data2['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 fan_off_testing_matrix[2]['speSend'] = '"' + parameter_tx + '"'
#
#                                 # step3 根据prv的paramter中第二行的RX填写第一行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 fan_off_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 fan_off_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 fan_off_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + '"'
#
#                                 # step3.1 根据prv的paramter中第二行的RX填写第三行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 fan_off_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 fan_off_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 fan_off_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#
#                                 # step4 根据prv的第二行LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                 lsl = data2['LSL']
#                                 usl = data2['USL']
#                                 unit = data2['Unit']
#                                 fan_off_testing_matrix[4]['LoLimit'] = lsl
#                                 fan_off_testing_matrix[4]['HiLimit'] = usl
#                                 fan_off_testing_matrix[4]['Unit'] = unit
#                                 # step5 根据prv中第二行的description 填写最后一行的speReference
#                                 match = re.search(r'check\s*byte\s*(\d+)\s*[-~]\s*(\d+)', data2['Description'],
#                                                   re.IGNORECASE)
#                                 if match:
#                                     byte_start = int(match.group(1))
#                                     byte_end = int(match.group(2))
#                                     print(
#                                         f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#                                     if byte_start == 0:
#                                         # check从头开始不需要标记位置 直接填写需要check的长度 byte_end 就是byte length
#                                         byte_length = byte_end
#                                         number = int(byte_length * 3 - 1)
#                                         fan_off_testing_matrix[4]['speReference'] = '"%s' + str(number) + '"'
#
#                                     else:
#                                         # first_number代表开始check的位置 second_number代表check的长度
#                                         first_number = byte_start * 3
#                                         byte_length = int(byte_end - byte_start + 1)
#                                         second_number = int(byte_length * 3 - 1)
#                                         fan_off_testing_matrix[4]['speReference'] = '"%s*' + str(
#                                             first_number) + '%s' + str(second_number) + '"'
#                                 else:
#                                     print("No match found.")
#
#                                 # step6 将第二行的test id 填写入最后一行的speResult
#                                 fan_off_testing_matrix[4]['speResult'] = fan_off_testing_matrix[3]['Test Id']
#
#                                 print("fan_off_testing_matrix： ", fan_off_testing_matrix)
#                                 # 存入字典
#                                 testing_matrix_dict['station'] = data['Station']
#                                 testing_matrix_dict['category'] = 'FAN OFF test'
#                                 testing_matrix_dict['testing_content'] = fan_off_testing_matrix
#                                 print("check one part testing_matrix_dict: ", testing_matrix_dict)
#                                 # 加入list
#                                 page_testing_matrix.append(testing_matrix_dict)
#                             except Exception as e:
#                                 print(f"❌ Error transfer prv content to testing matrix in fan off category: {e}")
#
#                         # 场景六：【FAN】 Fan PWM 100%
#                         if '[FAN]' in data['Program'] and 'PWM' in data['Program']:
#                             data2 = in_use_data_list[1]
#                             fan_pwm_testing_matrix = [
#                                 {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC",
#                                  "Description": "Set FAN test PWM 100%",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "10",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".02", "Input Message": "sUTIL_Delay",
#                                  "Description": "Waiting 1s for DUT FAN test PWM 100%",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "0",
#                                  "HiLimit": "0", "Unit": "Dec", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "1",
#                                  "speTimeOut": "", "speDelay": "1000", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "DELAY_DESCRIPTION {Waiting DUT Ready}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".03", "Input Message": "sDoIP_SND_RCV_RC",
#                                  "Description": "Read DUT FAN test PWM 100% ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": ".04", "Input Message": "sUTIL_VerifyResult",
#                                  "Description": "Parse DUT FAN test PWM 100% ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "%s*",
#                                  "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                  "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "Rm_DoIP_Param1", "speParamString": "RESULT_TYPE {String}",
#                                  "speResultString": "", "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                 {"Test Id": "", "Input Message": "sUTIL_ParseTASResultAsHex",
#                                  "Description": "Check DUT FAN test PWM 100% ADC Status",
#                                  "Bitmap": "", "speImg_xy": "",
#                                  "LoLimit": "",
#                                  "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "", "speFailGo": "",
#                                  "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                  "speRepeat": "",
#                                  "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                  "speCycleCounter": "",
#                                  "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                  "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                             print("***场景六匹配：FAN PWM ***")
#                             try:
#                                 # step1 根据prv的tasid填写所有test id
#                                 tas_id = data['TAS ID']
#                                 tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                 for index, matrix_row in enumerate(fan_pwm_testing_matrix):
#                                     if index == 4:
#                                         matrix_row['Test Id'] = tas_id
#                                     else:
#                                         matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#
#                                 # step2 根据prv的paramter中第一行的TX填写第一行的speSend 并且需要替换xx
#                                 parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 match_replace_xx_value = re.search(r'set\s*byte0:\s*0x([0-9a-fA-F]+)',
#                                                                    data['Description'], re.IGNORECASE)
#                                 if match_replace_xx_value:
#                                     replace_xx_value = str(match_replace_xx_value.group(1))
#                                     fan_pwm_testing_matrix[0]['speSend'] = '"' + parameter_tx.replace('xx',
#                                                                                                       replace_xx_value) + '"'
#                                 else:
#                                     print("No match_replace_xx_value found in FAN PWM.")
#
#                                 # step2.1 根据prv的paramter中第二行的TX填写第三行的speSend
#                                 parameter_tx = data2['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                 fan_pwm_testing_matrix[2]['speSend'] = '"' + parameter_tx + '"'
#
#                                 # step3 根据prv的paramter中第二行的RX填写第一行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 fan_pwm_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 fan_pwm_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 fan_pwm_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + '"'
#
#                                 # step3.1 根据prv的paramter中第二行的RX填写第三行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                 parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                           '').replace(
#                                     'zz', '').strip()
#                                 parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                 parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                 fan_pwm_testing_matrix[0]['LoLimit'] = parameter_rx
#                                 fan_pwm_testing_matrix[0]['HiLimit'] = parameter_rx
#                                 fan_pwm_testing_matrix[0]['speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#
#                                 # step4 根据prv的第二行LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                 lsl = data2['LSL']
#                                 usl = data2['USL']
#                                 unit = data2['Unit']
#                                 fan_pwm_testing_matrix[4]['LoLimit'] = lsl
#                                 fan_pwm_testing_matrix[4]['HiLimit'] = usl
#                                 fan_pwm_testing_matrix[4]['Unit'] = unit
#                                 # step5 根据prv中第二行的description 填写最后一行的speReference
#                                 match = re.search(r'check\s*byte\s*(\d+)\s*[-~]\s*(\d+)', data2['Description'],
#                                                   re.IGNORECASE)
#                                 if match:
#                                     byte_start = int(match.group(1))
#                                     byte_end = int(match.group(2))
#                                     print(
#                                         f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#                                     if byte_start == 0:
#                                         # check从头开始不需要标记位置 直接填写需要check的长度 byte_end 就是byte length
#                                         byte_length = byte_end
#                                         number = int(byte_length * 3 - 1)
#                                         fan_pwm_testing_matrix[4]['speReference'] = '"%s' + str(number) + '"'
#
#                                     else:
#                                         # first_number代表开始check的位置 second_number代表check的长度
#                                         first_number = byte_start * 3
#                                         byte_length = int(byte_end - byte_start + 1)
#                                         second_number = int(byte_length * 3 - 1)
#                                         fan_pwm_testing_matrix[4]['speReference'] = '"%s*' + str(
#                                             first_number) + '%s' + str(second_number) + '"'
#                                 else:
#                                     print("No match found.")
#
#                                 # step6 将第二行的test id 填写入最后一行的speResult
#                                 fan_pwm_testing_matrix[4]['speResult'] = fan_pwm_testing_matrix[3]['Test Id']
#
#                                 print("fan_pwm_testing_matrix： ", fan_pwm_testing_matrix)
#                                 # 存入字典
#                                 testing_matrix_dict['station'] = data['Station']
#                                 testing_matrix_dict['category'] = 'FAN test PWM 100%'
#                                 testing_matrix_dict['testing_content'] = fan_pwm_testing_matrix
#                                 print("check one part testing_matrix_dict: ", testing_matrix_dict)
#                                 # 加入list
#                                 page_testing_matrix.append(testing_matrix_dict)
#                             except Exception as e:
#                                 print(f"❌ Error transfer prv content to testing matrix in fan pwm category: {e}")
#
#                         # 场景七：【Video Out GPIO】 场景八 【Video In GPIO]
#                         if '[Video Out GPIO]' in data['Program'] or '[Video In GPIO]' in data['Program']:
#                             if len(in_use_data_list) == 2:
#                                 if data['Action'] != 'TASORDER':
#                                     data2 = in_use_data_list[1]
#                                     video_out_gpio_testing_matrix = [
#                                         {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "",
#                                          "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "10",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "",
#                                          "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": ".02", "Input Message": "sUTIL_VerifyResult", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "%s*",
#                                          "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "Rm_DoIP_Param1",
#                                          "speParamString": "RESULT_TYPE{String}",
#                                          "speResultString": "", "speBlockCode": "", "PromptMessage": "",
#                                          "LastColumn": ""},
#                                         {"Test Id": ".03", "Input Message": "sUTIL_ParseTASResultAsHex",
#                                          "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "",
#                                          "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "",
#                                          "speResultString": "", "speBlockCode": "", "PromptMessage": "",
#                                          "LastColumn": ""},
#                                         {"Test Id": ".04", "Input Message": "sUTIL_Delay", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "0",
#                                          "HiLimit": "0", "Unit": "Dec", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "1",
#                                          "speTimeOut": "", "speDelay": "1000", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "",
#                                          "speParamString": "DELAY_DESCRIPTION {Waiting DUT Ready}",
#                                          "speResultString": "", "speBlockCode": "", "PromptMessage": "",
#                                          "LastColumn": ""},
#                                         {"Test Id": ".05", "Input Message": "sDoIP_SND_RCV_RC", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "",
#                                          "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "10",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": ".06", "Input Message": "sUTIL_VerifyResult", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "%s*",
#                                          "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "Rm_DoIP_Param1",
#                                          "speParamString": "RESULT_TYPE {String}",
#                                          "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": ".07", "Input Message": "sUTIL_ParseTASResultAsHex",
#                                          "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "0",
#                                          "HiLimit": "0", "Unit": "Hex", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": "", "Input Message": "sUTIL_ParseTASResult", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "",
#                                          "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "RESULT_MANIPULATION{mPas_AsciiHexToString}",
#                                          "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                                     print("***场景七匹配：VIDEO OUT PGIO  or  场景八匹配：VIDEO OUT PGIO***")
#                                     try:
#                                         # step1 根据prv的tasid填写所有test id
#                                         tas_id = data['TAS ID']
#                                         tef_id = data['TAS ID'].replace('TAS', 'TEF')
#                                         for index, matrix_row in enumerate(video_out_gpio_testing_matrix):
#                                             if index == 7:
#                                                 matrix_row['Test Id'] = tas_id
#                                             else:
#                                                 matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#
#                                         # step2 根据prv的paramter中第一行的TX填写第一行的speSend
#                                         parameter_tx = data['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                         video_out_gpio_testing_matrix[0]['speSend'] = '"' + parameter_tx + '"'
#
#                                         # step2.1 根据prv的paramter中第二行的TX填写第五行的speSend
#                                         parameter_tx = data2['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                         video_out_gpio_testing_matrix[4]['speSend'] = '"' + parameter_tx + '"'
#
#                                         # step3 根据prv的paramter中第一行的RX填写第一行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                         parameter_rx = data['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                                   '').replace(
#                                             'zz', '').strip()
#                                         parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                         parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                         video_out_gpio_testing_matrix[0]['LoLimit'] = parameter_rx
#                                         video_out_gpio_testing_matrix[0]['HiLimit'] = parameter_rx
#                                         video_out_gpio_testing_matrix[0][
#                                             'speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#
#                                         # step3.1 根据prv的paramter中第二行的RX填写第五行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                         parameter_rx = data2['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                                    '').replace(
#                                             'zz', '').strip()
#                                         parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                         parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                         video_out_gpio_testing_matrix[4]['LoLimit'] = parameter_rx
#                                         video_out_gpio_testing_matrix[4]['HiLimit'] = parameter_rx
#                                         video_out_gpio_testing_matrix[4][
#                                             'speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#
#                                         # step4 根据prv的第二行LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                         lsl = data2['LSL']
#                                         usl = data2['USL']
#                                         unit = data2['Unit']
#                                         video_out_gpio_testing_matrix[7]['LoLimit'] = lsl
#                                         video_out_gpio_testing_matrix[7]['HiLimit'] = usl
#                                         video_out_gpio_testing_matrix[7]['Unit'] = unit
#
#                                         # step4.1 根据prv的第一行LSL USL Unit 填充第三行的 LoLimit HiLimit Unit
#                                         lsl = data['LSL']
#                                         usl = data['USL']
#                                         unit = data['Unit']
#                                         video_out_gpio_testing_matrix[2]['LoLimit'] = lsl
#                                         video_out_gpio_testing_matrix[2]['HiLimit'] = usl
#                                         video_out_gpio_testing_matrix[2]['Unit'] = unit
#
#                                         # step5 根据prv中第一行的description 填写第三行的speReference
#                                         match = re.search(r'check\s*byte\s*(\d+)(?:\s*[-~]\s*(\d+))?',
#                                                           data['Description'],
#                                                           re.IGNORECASE)
#                                         if match:
#                                             byte_start = int(match.group(1))
#                                             byte_end = int(match.group(2)) if match.group(2) else None
#                                             # 在有区间的情况下
#                                             if byte_end is not None:
#                                                 print(
#                                                     f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#                                                 if byte_start == 0:
#                                                     # check从头开始不需要标记位置 直接填写需要check的长度 byte_end 就是byte length
#                                                     byte_length = byte_end
#                                                     number = int(byte_length * 3 - 1)
#                                                     video_out_gpio_testing_matrix[2]['speReference'] = '"%s' + str(
#                                                         number) + '"'
#
#                                                 else:
#                                                     # first_number代表开始check的位置 second_number代表check的长度
#                                                     first_number = byte_start * 3
#                                                     byte_length = int(byte_end - byte_start + 1)
#                                                     second_number = int(byte_length * 3 - 1)
#                                                     video_out_gpio_testing_matrix[2]['speReference'] = '"%s*' + str(
#                                                         first_number) + '%s' + str(second_number) + '"'
#                                             # 没有区间单个的情况下 check都是2位str 起始位会不一样
#                                             else:
#                                                 if byte_start == 0:
#                                                     video_out_gpio_testing_matrix[2]['speReference'] = '"%s2"'
#
#                                                 else:
#                                                     first_number = byte_start * 3
#                                                     video_out_gpio_testing_matrix[2]['speReference'] = '"%s*' + str(
#                                                         first_number) + '%s2"'
#
#                                         else:
#                                             print("No match found.")
#
#                                         # step5.1 根据prv中第二行的description 填写第七第八行的speReference
#                                         row_index = 6  # 从第7行开始（Python索引）
#                                         for second_row_description_sentence in data2['Description'].split('\n'):
#                                             match = re.search(r'check\s*byte\s*(\d+)(?:\s*[-~]\s*(\d+))?',
#                                                               second_row_description_sentence, re.IGNORECASE)
#                                             if match:
#                                                 print('match description: ', second_row_description_sentence)
#                                                 byte_start = int(match.group(1))
#                                                 byte_end = int(match.group(2)) if match.group(2) else None
#
#                                                 if byte_end is not None:
#                                                     print(
#                                                         f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#
#                                                     if byte_start == 0:
#                                                         # 从头开始 check，长度就是 byte_end
#                                                         byte_length = byte_end
#                                                         number = int(byte_length * 3 - 1)
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s' + str(
#                                                             number) + '"'
#                                                     else:
#                                                         first_number = byte_start * 3
#                                                         byte_length = byte_end - byte_start + 1
#                                                         second_number = int(byte_length * 3 - 1)
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s*' + str(
#                                                             first_number) + '%s' + str(second_number) + '"'
#                                                 else:
#                                                     print(f"Single byte check at: {byte_start}")
#
#                                                     if byte_start == 0:
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s2"'
#                                                     else:
#                                                         first_number = byte_start * 3
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s*' + str(
#                                                             first_number) + '%s2"'
#
#                                                 row_index += 1
#                                                 if row_index > 7:
#                                                     break
#                                             else:
#                                                 print('No match description: ', second_row_description_sentence)
#                                                 print("No match found.")
#
#                                         # step6 将第二行的test id 填写入第三行的speResult
#                                         video_out_gpio_testing_matrix[2]['speResult'] = \
#                                         video_out_gpio_testing_matrix[1][
#                                             'Test Id']
#
#                                         # step6.1 将第六行的test id 填写入第七八行的speResult
#                                         video_out_gpio_testing_matrix[6]['speResult'] = \
#                                         video_out_gpio_testing_matrix[5][
#                                             'Test Id']
#                                         video_out_gpio_testing_matrix[7]['speResult'] = \
#                                         video_out_gpio_testing_matrix[5][
#                                             'Test Id']
#
#                                         # step7 description
#                                         program_name = data['Program'].replace('[Video Out GPIO]', '').replace(
#                                             '[Video In GPIO]', '').strip()
#                                         video_out_gpio_testing_matrix[0]['Description'] = "Trigger DUT " + program_name
#                                         video_out_gpio_testing_matrix[1][
#                                             'Description'] = "Parse DUT " + program_name + " trigger status"
#                                         video_out_gpio_testing_matrix[2][
#                                             'Description'] = "Check DUT " + program_name + " trigger status"
#                                         video_out_gpio_testing_matrix[3][
#                                             'Description'] = "Waiting 1s for DUT " + program_name + " status"
#                                         video_out_gpio_testing_matrix[4][
#                                             'Description'] = "Get " + program_name + " Ser results"
#                                         video_out_gpio_testing_matrix[5][
#                                             'Description'] = "Parse " + program_name + " Ser results"
#                                         video_out_gpio_testing_matrix[6][
#                                             'Description'] = "Check " + program_name + " Ser status"
#                                         video_out_gpio_testing_matrix[7][
#                                             'Description'] = "Check " + program_name + " Ser name"
#
#                                         print("video_out/in_gpio_testing_matrix： ", video_out_gpio_testing_matrix)
#                                         # 存入字典
#                                         testing_matrix_dict['station'] = data['Station']
#                                         testing_matrix_dict['category'] = program_name
#                                         testing_matrix_dict['testing_content'] = video_out_gpio_testing_matrix
#                                         # 加入list
#                                         page_testing_matrix.append(testing_matrix_dict)
#                                     except Exception as e:
#                                         print(
#                                             f"❌ Error transfer prv content to testing matrix in video out/in gpio category: {e}")
#                                 else:
#                                     data2 = in_use_data_list[1]
#                                     video_out_gpio_testing_matrix = [
#                                         {"Test Id": ".01", "Input Message": "sDoIP_SND_RCV_RC", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "",
#                                          "HiLimit": "", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "10",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": ".02", "Input Message": "sUTIL_VerifyResult", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "%s*",
#                                          "HiLimit": "%s*", "Unit": "String", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "Rm_DoIP_Param1",
#                                          "speParamString": "RESULT_TYPE {String}",
#                                          "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": ".03", "Input Message": "sUTIL_ParseTASResultAsHex",
#                                          "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "0",
#                                          "HiLimit": "0", "Unit": "Hex", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "", "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""},
#                                         {"Test Id": "", "Input Message": "sUTIL_ParseTASResult", "Description": "",
#                                          "Bitmap": "", "speImg_xy": "",
#                                          "LoLimit": "",
#                                          "HiLimit": "", "Unit": "", "speJumpPass": "", "speJumpFail": "",
#                                          "speFailGo": "",
#                                          "speReference": "", "speSend": "", "speReceive": "", "speBusTimeOut": "",
#                                          "speRepeat": "",
#                                          "speTimeOut": "", "speDelay": "", "speDelayRepeat": "", "speResult": "",
#                                          "speCycleCounter": "",
#                                          "speRegisterTransfer": "", "speParamString": "RESULT_MANIPULATION{mPas_AsciiHexToString}",
#                                          "speResultString": "",
#                                          "speBlockCode": "", "PromptMessage": "", "LastColumn": ""}]
#                                     print("***场景七匹配：VIDEO OUT PGIO  or  场景八匹配：VIDEO OUT PGIO WITH FOLLOW TASORDER***")
#                                     try:
#                                         # step1 根据prv的tasid填写所有test id
#                                         tas_id = data2['TAS ID']
#                                         tef_id = data2['TAS ID'].replace('TAS', 'TEF')
#                                         for index, matrix_row in enumerate(video_out_gpio_testing_matrix):
#                                             if index == 3:
#                                                 matrix_row['Test Id'] = tas_id
#                                             else:
#                                                 matrix_row['Test Id'] = tef_id + matrix_row['Test Id']
#
#                                         # step2.1 根据prv的paramter中第二行的TX填写第五行的speSend
#                                         parameter_tx = data2['Parameter'].split('|')[0].replace('TX:', '').strip()
#                                         video_out_gpio_testing_matrix[0]['speSend'] = '"' + parameter_tx + '"'
#
#                                         # step3.1 根据prv的paramter中第二行的RX填写第五行的LoLimit和HiLimit, 计算zz 前内容的长度填充到speReceive, “%z长度 %z
#                                         parameter_rx = data2['Parameter'].split('|')[1].replace('RX:', '').replace('ZZ',
#                                                                                                                    '').replace(
#                                             'zz', '').strip()
#                                         parameter_rx_without_space = parameter_rx.replace(" ", "")
#                                         parameter_rx_length_str = str(int(len(parameter_rx_without_space) / 2))
#                                         video_out_gpio_testing_matrix[0]['LoLimit'] = parameter_rx
#                                         video_out_gpio_testing_matrix[0]['HiLimit'] = parameter_rx
#                                         video_out_gpio_testing_matrix[0][
#                                             'speReceive'] = '"%z' + parameter_rx_length_str + ' %z"'
#
#                                         # step4 根据prv的第二行LSL USL Unit 填充最后一行的 LoLimit HiLimit Unit
#                                         lsl = data2['LSL']
#                                         usl = data2['USL']
#                                         unit = data2['Unit']
#                                         video_out_gpio_testing_matrix[3]['LoLimit'] = lsl
#                                         video_out_gpio_testing_matrix[3]['HiLimit'] = usl
#                                         video_out_gpio_testing_matrix[3]['Unit'] = unit
#
#                                         # step5.1 根据prv中第二行的description 填写第七第八行的speReference
#                                         row_index = 2  # 从第7行开始（Python索引）
#                                         for second_row_description_sentence in data2['Description'].split('\n'):
#                                             match = re.search(r'check\s*byte\s*(\d+)(?:\s*[-~]\s*(\d+))?',
#                                                               second_row_description_sentence, re.IGNORECASE)
#                                             if match:
#                                                 print('match description: ', second_row_description_sentence)
#                                                 byte_start = int(match.group(1))
#                                                 byte_end = int(match.group(2)) if match.group(2) else None
#
#                                                 if byte_end is not None:
#                                                     print(
#                                                         f"Byte range: {byte_start} to {byte_end}, and length: {byte_end - byte_start + 1}")
#
#                                                     if byte_start == 0:
#                                                         # 从头开始 check，长度就是 byte_end
#                                                         byte_length = byte_end
#                                                         number = int(byte_length * 3 - 1)
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s' + str(
#                                                             number) + '"'
#                                                     else:
#                                                         first_number = byte_start * 3
#                                                         byte_length = byte_end - byte_start + 1
#                                                         second_number = int(byte_length * 3 - 1)
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s*' + str(
#                                                             first_number) + '%s' + str(second_number) + '"'
#                                                 else:
#                                                     print(f"Single byte check at: {byte_start}")
#
#                                                     if byte_start == 0:
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s2"'
#                                                     else:
#                                                         first_number = byte_start * 3
#                                                         video_out_gpio_testing_matrix[row_index][
#                                                             'speReference'] = '"%s*' + str(
#                                                             first_number) + '%s2"'
#
#                                                 row_index += 1
#                                                 if row_index > 3:
#                                                     break
#                                             else:
#                                                 print('No match description: ', second_row_description_sentence)
#                                                 print("No match found.")
#
#                                         # step6.1 将第六行的test id 填写入第七八行的speResult
#                                         video_out_gpio_testing_matrix[2]['speResult'] = \
#                                         video_out_gpio_testing_matrix[1][
#                                             'Test Id']
#                                         video_out_gpio_testing_matrix[3]['speResult'] = \
#                                         video_out_gpio_testing_matrix[1][
#                                             'Test Id']
#
#                                         # step7 description
#                                         program_name = data['Program'].replace('[Video Out GPIO]', '').replace(
#                                             '[Video In GPIO]', '').strip()
#
#                                         video_out_gpio_testing_matrix[0][
#                                             'Description'] = "Get " + program_name + " Ser results"
#                                         video_out_gpio_testing_matrix[1][
#                                             'Description'] = "Parse " + program_name + " Ser results"
#                                         video_out_gpio_testing_matrix[2][
#                                             'Description'] = "Check " + program_name + " Ser status"
#                                         video_out_gpio_testing_matrix[3][
#                                             'Description'] = "Check " + program_name + " Ser name"
#
#                                         print("video_out/in_gpio_testing_matrix with follow tasorder： ",
#                                               video_out_gpio_testing_matrix)
#                                         # 存入字典
#                                         testing_matrix_dict['station'] = data['Station']
#                                         testing_matrix_dict['category'] = program_name
#                                         testing_matrix_dict['testing_content'] = video_out_gpio_testing_matrix
#                                         print("check one part testing_matrix_dict: ", testing_matrix_dict)
#                                         # 加入list
#                                         page_testing_matrix.append(testing_matrix_dict)
#                                     except Exception as e:
#                                         print(
#                                             f"❌ Error transfer prv content to testing matrix in video out/in gpio with follow tasorder category: {e}")
#
#                             else:
#                                 print(
#                                     f"⚠ Parsing ERROR UNDER VIDEO IN/OUT GPIO category,ERROR CAUSING BY DIDN'T GET TWO ROWS")
#
#
#
#                     # 异步传输每个image和content
#                     result_dict = {'page_image_path': image_path, 'page_image_content': data_list,
#                                    'testing_matrix': page_testing_matrix}
#                     print("！！！ 发送给前端的结果： ", result_dict)
#                     prv_message = {'connectionID': task_id,
#                                    'category': 'prv_image', 'from': '', 'to': '',
#                                    'message': json.dumps(result_dict, ensure_ascii=False),
#                                    'remarks': json.dumps({'paragraph_start': 1,
#                                                           'response_end': 0})}
#                     try:
#                         asyncio.run(websocket_client(prv_message))
#                     except:
#                         time.sleep(0.5)
#                         asyncio.run(websocket_client(prv_message))
#
#                     # ⏬ 插入占行信息 + 扩展真实内容
#                     def extend_with_label(records, label, new_data):
#                         if new_data:
#                             records.append({'Test Id': label})
#                             records.extend(new_data)
#
#                     # 不为空再记录 后续存到excel
#                     if page_testing_matrix:
#                         for testing_group in page_testing_matrix:
#                             if testing_group['station'] == 'FCT50':
#                                 label = testing_group['category']
#                                 extend_with_label(records, label, testing_group['testing_content'])
#                 else:
#                     print(f"⚠️ 空内容跳过解析：{image_path}")
#
#             except Exception as e:
#                 print(f"❌ Error parsing {image_path}: {e}")
#                 # print(f"⚠️ 原始返回内容为： {repr(json_str)}")
#
#
#     # 结果生成excel存储到github
#     if records:
#         print("✔ Excel has content to download!!!!")
#         all_keys = set()
#         for r in records:
#             all_keys.update(r.keys())
#
#         # 字段统一化处理
#         normalized_records = []
#         for r in records:
#             normalized_records.append({key: r.get(key, None) for key in all_keys})
#
#         df = pd.DataFrame(records)
#         excel_name = os.path.normpath(pdf_path).split('/')[-1].replace('.pdf','.xlsx')
#         print("download excel name: ", excel_name)
#         output_path = os.path.join(root_directory, 'prv_pdf_images', 'FCT50_' + excel_name)
#         df.to_excel(output_path, index=False, sheet_name='spe')
#
#         # 打开刚写好的 Excel 文件
#         wb = load_workbook(output_path)
#         ws = wb['spe']
#         max_col = ws.max_column
#
#         # 获取 "Test Id" 所在的列号（假设第一行为表头）
#         header = [cell.value for cell in ws[1]]
#         try:
#             test_id_col_idx = header.index("Test Id") + 1  # openpyxl 列号从 1 开始
#         except ValueError:
#             raise Exception("❌ 没有找到列名 'Test Id'，请确认表头是否正确")
#
#         # 遍历每一行（从第2行开始）
#         for row in range(2, ws.max_row + 1):
#             test_id_value = ws.cell(row=row, column=test_id_col_idx).value
#             # 除 Test Id 外其它列是否全空？
#             others_empty = all(
#                 ws.cell(row=row, column=col).value in [None, ""]
#                 for col in range(1, max_col + 1) if col != test_id_col_idx
#             )
#
#             if test_id_value and others_empty:
#                 # 合并整行（从A列到最后一列）
#                 ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
#                 # 设置居中
#                 ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
#
#         # 保存
#         wb.save(output_path)
#         download_excel_path_dict['FCT50_' + excel_name] = output_path
#
#     # 异步传输downloadexcel地址信息
#     download_message = {'connectionID': task_id,
#                    'category': 'downloadfile', 'from': '', 'to': '',
#                    'message': json.dumps(download_excel_path_dict, ensure_ascii=False),
#                     'remarks': json.dumps({'paragraph_start': 1,
#                                         'response_end': 0})}
#     try:
#         asyncio.run(websocket_client(download_message))
#     except:
#         time.sleep(0.5)
#         asyncio.run(websocket_client(download_message))
#
#     # 异步传输结束信号空字符串
#     prv_message = {'connectionID': task_id,
#                    'category': 'text', 'from': '', 'to': '',
#                    'message': '',
#                    'remarks': json.dumps({'paragraph_start': 0,
#                                           'response_end': 1})}
#     try:
#         asyncio.run(websocket_client(prv_message))
#     except:
#         time.sleep(0.5)
#         asyncio.run(websocket_client(prv_message))



###################################################################################################

                             ##       WUJ AUTO PRV  PHASE 2      ##

###################################################################################################

def auto_prv_improve_by_action(task_id,pdf_path, excel_path):

    # 先读取 excel content
    # response = requests.get(excel_path, verify=False)
    # response.raise_for_status()
    # file_like = BytesIO(response.content)
    excel_table = pd.read_excel(excel_path)



    # original_pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
    # print("save original_pdf_name", original_pdf_name)
    download_excel_path_dict = {}
    # def clean_llm_json(json_str):
    #     # 去掉前后的 markdown 代码块标识符 ```json ... ```
    #     if json_str.startswith("```"):
    #         json_str = re.sub(r"^```(?:json)?\n", "", json_str)
    #         json_str = re.sub(r"\n```$", "", json_str)
    #     return json_str.strip()



    def clean_llm_json(json_str: str) -> str:
        """
        清理 LLM 返回的 JSON 字符串：
        1. 去掉 ```json ... ``` 包裹
        2. 修复被截断的长字符串字段（特别是最后一个 Parameter）
        3. 补齐缺失的括号
        4. 返回可解析的 JSON 字符串
        """
        # ========== 1. 去掉 Markdown 包裹 ==========
        if json_str.startswith("```"):
            json_str = re.sub(r"^```(?:json)?\n", "", json_str)
            json_str = re.sub(r"\n```$", "", json_str)

        json_str = json_str.strip()

        # ========== 2. 尝试直接解析 ==========
        try:
            json.loads(json_str)
            return json_str
        except json.JSONDecodeError as e:
            print(f"❌ JSON 初次解析失败: {e}")

        fixed = json_str

        # ========== 3. 修复被截断的字符串 ==========
        # 找到最后一个 "Parameter": 开头的字段
        param_match = re.search(r'"Parameter"\s*:\s*"(.*)', fixed, re.DOTALL)
        if param_match:
            before = fixed[:param_match.end(1)]
            after = fixed[param_match.end(1):]

            # 如果字符串在中途被截断（没有找到正确的结束引号）
            if not after.strip().startswith('"'):
                print("⚠️ 检测到 Parameter 字段被截断 → 强制闭合引号")
                fixed = before + '"'  # 给 Parameter 强制闭合引号
                # 如果对象没收尾，加上 }
                if not fixed.strip().endswith("}"):
                    fixed += "}"
                # 如果数组没收尾，加上 ]
                if fixed.strip().startswith("[") and not fixed.strip().endswith("]"):
                    fixed += "]"

        # ========== 4. 再次尝试解析 ==========
        try:
            json.loads(fixed)
            print("✅ JSON 修复成功")
            return fixed
        except json.JSONDecodeError as e:
            snippet = fixed[e.pos - 30:e.pos + 30] if hasattr(e, "pos") else fixed[:60]
            raise ValueError(
                f"❌ JSON 修复失败: {e}\n"
                f"➡️ 错误附近内容: {snippet}\n"
                f"➡️ 修复后长度: {len(fixed)} 字符"
            )

    # def clean_llm_json(json_str: str) -> str:
    #     """
    #     清理 LLM 返回的 JSON 字符串：
    #     1. 去掉 ```json ... ``` 包裹
    #     2. 检查并补全可能缺失的引号/括号
    #     3. 返回可解析的 JSON 字符串
    #     """
    #     # ========== 1. 去掉 Markdown 包裹 ==========
    #     if json_str.startswith("```"):
    #         json_str = re.sub(r"^```(?:json)?\n", "", json_str)
    #         json_str = re.sub(r"\n```$", "", json_str)
    #
    #     json_str = json_str.strip()
    #
    #     # ========== 2. 尝试直接解析 ==========
    #     try:
    #         json.loads(json_str)
    #         return json_str
    #     except json.JSONDecodeError as e:
    #         print(f"❌ JSON 初次解析失败: {e}")+

    #
    #     # ========== 3. 自动诊断 ==========
    #     lines = json_str.splitlines()
    #     last_line = lines[-1] if lines else ""
    #     print("🔍 JSON 最后一行:", last_line)
    #
    #     fixed = json_str
    #
    #     # 情况 A: 字符串缺少收尾引号
    #     if not fixed.endswith(("]", "}")):
    #         last_quote = fixed.rfind('"')
    #         last_colon = fixed.rfind(':')
    #         if last_colon > last_quote:
    #             print("⚠️ 检测到最后一个字段缺少引号 → 自动补齐 '\"'")
    #             fixed += '"'
    #
    #         # 情况 B: 缺少整体收尾括号
    #         if fixed.strip().startswith("[") and not fixed.strip().endswith("]"):
    #             print("⚠️ 检测到数组缺少结尾 ']' → 自动补齐")
    #             fixed += "]"
    #         elif fixed.strip().startswith("{") and not fixed.strip().endswith("}"):
    #             print("⚠️ 检测到对象缺少结尾 '}' → 自动补齐")
    #             fixed += "}"
    #
    #     # ========== 4. 再次尝试解析 ==========
    #     try:
    #         json.loads(fixed)
    #         print("✅ JSON 修复成功")
    #         return fixed
    #     except json.JSONDecodeError as e:
    #         snippet = fixed[e.pos - 30:e.pos + 30] if hasattr(e, "pos") else fixed[:60]
    #         raise ValueError(
    #             f"❌ JSON 修复失败: {e}\n"
    #             f"➡️ 错误附近内容: {snippet}\n"
    #             f"➡️ 修复后长度: {len(fixed)} 字符"
    #         )

        # GPT-4o 调用

    # def preprocess_image(image_bytes, debug_dir=None):
    #     """
    #     专门针对技术文档的增强图像预处理，提高OCR识别精度
    #
    #     参数:
    #         image_bytes: 原始图片字节流
    #         debug_dir: 调试目录，用于保存中间处理步骤的图像
    #
    #     返回:
    #         处理后的图像字节流
    #     """
    #
    #     def save_debug(img, name):
    #         """保存调试图像到指定目录"""
    #         if debug_dir:
    #             os.makedirs(debug_dir, exist_ok=True)
    #             cv2.imwrite(os.path.join(debug_dir, f"{name}.png"), img)
    #
    #     # 1. 读取图像数据
    #     img_array = np.frombuffer(image_bytes, np.uint8)
    #     img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    #     save_debug(img, "0_原始图像")
    #
    #     # 2. 转换为灰度图像进行处理
    #     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    #     save_debug(gray, "1_灰度图像")
    #
    #     # 3. 应用高斯模糊减少噪声
    #     blurred = cv2.GaussianBlur(gray, (3, 3), 0)
    #     save_debug(blurred, "2_模糊处理")
    #
    #     # 4. 增强对比度 - 使用更小的瓦片尺寸提高局部对比度
    #     clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(4, 4))
    #     enhanced = clahe.apply(blurred)
    #     save_debug(enhanced, "3_对比度增强")
    #
    #     # 5. 形态学操作清理文本
    #     kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    #     morphed = cv2.morphologyEx(enhanced, cv2.MORPH_CLOSE, kernel)
    #     save_debug(morphed, "4_形态学处理")
    #
    #     # 6. 自适应阈值处理 - 优化参数以提高文本清晰度
    #     binary = cv2.adaptiveThreshold(
    #         morphed, 255,
    #         cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
    #         cv2.THRESH_BINARY,
    #         15, 8  # 更小的块大小和C值，提高文本清晰度
    #     )
    #     save_debug(binary, "5_自适应阈值")
    #
    #     # 7. 移除小噪点
    #     kernel = np.ones((2, 2), np.uint8)
    #     cleaned = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
    #     save_debug(cleaned, "6_噪点清理")
    #
    #     # 8. 显著放大图像以提高OCR识别率
    #     scale_factor = 3  # 增加放大倍数至3倍
    #     height, width = cleaned.shape
    #     resized = cv2.resize(cleaned, (width * scale_factor, height * scale_factor),
    #                          interpolation=cv2.INTER_CUBIC)
    #     save_debug(resized, "7_放大处理")
    #
    #     # 9. 转换回字节流格式
    #     success, buf = cv2.imencode(".png", resized)
    #     if not success:
    #         raise Exception("图像预处理失败")
    #     return buf.tobytes()

    def create_enhanced_unified_prompt():
        """
        完善的统一提示模板
        结合字符识别纠错、表格完整性检查和结构化输出要求
        """
        return """你将看到一张包含EOL测试步骤的技术文档图片。请严格按照以下要求完整、准确地提取所有信息：

    📋 图片内容结构：
    图片包含以下元素：
    - TAS编码：以"TAS"开头并以五位数字结尾（例如TAS20678）
    - Station：TAS编码旁的灰色小字体站点信息（例如 FCT50 / FCT60 /CUD50/FCT42)
    - Program：TAS编码和Station下一行的程序信息 
    - 数据表格：包含以下列的无框表格
      * Action（操作类型：DoIP_COM、SWITCH、DCV、DCI、RESISTANCE、MATH等）
      * LSL（下限值）
      * USL（上限值）
      * Unit（单位：Hex、Dec等）
      * Circuit（电路信息）
      * Parameter（参数：通常包含TX:、RX:格式的十六进制数据）
      * Description（详细描述）
    
    ⚠️ 表格列对齐关键规则：
    - 每行数据必须严格按照 Action → LSL → USL → Unit → Circuit → Parameter → Description 的列顺序对应
    - 即使某列内容为空，也要保持列位置不变，用空字符填充
    - 单元格会包含多行文本，不要拆分到其他列

    🎯 表格完整性要求（关键）：
    1. **逐行扫描原则**：从表格第一行到最后一行，按顺序检查每一行数据
    2. **行数完整验证**：提取前先目视统计表格总行数，确保提取的JSON对象数量与实际行数完全一致
    3. **重点行识别**：
       - 带绿色小箭头(→)的行："Arrow Tag"填"1"，这些行特别重要
       - Action为MATH的行：容易被忽略，必须识别
       - Action为SWITCH的行：容易被忽略，必须识别
       - 表格头下一行：经常被遗漏，特别关注
       - 表格最后几行：经常被遗漏，特别关注
    4. **边界清晰识别**：每行都有明确边界，即使行高不一致也要完整提取
    5. **列内容不要串列**：每列都有明确的边界，列内内容存在换行，请不要丢失或者填充到其它空白或非空白列
    

    ⚠️ 字符识别纠错规则（防止OCR错误）：

    关键字符对照表：
    - "zz" ≠ "22"：小写字母z，绝不是数字22
    - "|" ≠ "1"：竖直管道符，绝不是数字1
    - "6E" ≠ "8E"：数字6加字母E，不是数字8
    - "0D" ≠ "00"：数字0加字母D，不是两个0
    - "8" ≠ "B"：数字8有封闭圆环，字母B有直线段
    - "5" ≠ "S"：数字5有棱角，字母S是弯曲的
    - "0" ≠ "O"：在十六进制中通常是数字0
    - "00" ≠ "02"：关键区分
      * "00"：两个完整的圆形，通常用于填充或空值
      * "02"：圆形0后跟尖角数字2
      
    重要：Parameter字段长度限制
    - 如果发现重复模式超过20次，立即停止并检查
    - 绝对不要生成无限重复的模式

    Parameter字段特殊验证：
    - 标准格式："TX:31 01 FE 12 00 | RX:71 01 FE 12 ZZ | TO:5000"
    - TX和RX用竖直条"|"分隔，不是数字"1"，不会有单个1出现
    - 十六进制值只包含0-9和A-F字符
    - "ZZ"在RX中很常见，不要误认为"22"
    - 如果看到重复数字（22、11、88），请验证是否应该是字母
    - 不会是无限
    
    字符识别检查清单：
    1. 检查所有"22"是否应该是"zz"
    2. 确认所有"|"没被识别为"1"
    3. 验证十六进制中的字母A-F
    4. 寻找TX:、RX:、0x等格式线索
    

    📤 输出格式要求：
    提取每行数据为一个JSON对象，组成数组：

    [
      {
        "Arrow Tag": "1或空字符串",
        "TAS ID": "TAS编码",
        "Station": "站点信息",
        "Program": "程序信息",
        "Action": "操作类型（举例：SWITCH/DoIP_COM/TASORDER/MATH/DCV_10/DCV_1/RESISTANCE/REMARK/DCI_S/LIN)",
        "LSL": "下限值（注意字符准确性，转折的内容不要丢失）",
        "USL": "上限值（注意字符准确性，转折的内容不要丢失）",
        "Unit": "单位",
        "Circuit": "电路",
        "Parameter": "参数内容（注意字符准确性，一定不要不限长度的重复,按照实际长度返回）",
        "Description": "描述内容（保持原格式和换行）"
      },
      // 表格中每一行都必须对应一个JSON对象
    ]

    🔍 提取前预检查：
    1. 先数一下表格有几行数据
    2. 识别哪些行有绿色箭头
    3. 确认所有Action类型都能看清
    4. 仔细观察列边界，确保每行数据正确对应到相应列
    5. 特别检查Parameter字段的实际长度，避免生成重复模式

    🚨 最终质量检查：
    - JSON数组长度 = 表格实际行数
    - 所有绿色箭头行都标记了"Arrow Tag":"1"
    - Parameter字段没有无限重复模式
    - Parameter字段中没有"22"误认为"zz"的情况
    - 所有"|"符号没被误认为"1"
    - Action包含了所有可见的类型（DoIP_COM、MATH等）
    - 空字段检查：不要因为某行多个字段为空就跳过该行，看到action一定是单独一行
    - 列对齐验证：每行的LSL、USL内容合理（数字、范围值、空值等）
    - 如果图片不符合EOL测试步骤格式，返回空数组[]
    
    ⚠️ 严格禁止：
    - 不要生成超过实际图像内容长度的数据
    - 不要重复生成相同的十六进制模式
    - 不要因为看到重复模式就无限延续
    - 如果不确定某个字段的确切长度，宁可截短也不要过度延伸
    - 只返回JSON数组，不要添加任何解释文字或注释

    ⚠️ 重要提醒：
    - 这是关键技术数据，字符错误会导致系统故障
    - 表格行遗漏会造成测试步骤缺失
    - 请务必完整、准确、仔细地提取每个细节
    - 只返回JSON数组，不要添加任何解释文字或注释"""

    # def create_character_mapping_prompt():
    #     """
    #     创建专门的字符识别纠正提示
    #     包含常见的OCR识别错误及其正确形式
    #     """
    #     return """
    #     ⚠️ 关键字符识别规则 - 请特别注意以下常见识别错误：
    #
    #     字符纠正对照表：
    #     - "zz" 应该是 "zz"（小写字母z，不是数字22）
    #     - "|" 是竖直管道符，不是数字"1"
    #     - "6E" 应该是 "6E"（数字6，不是数字8）
    #     - "0D" 应该是 "0D"（字母D，不是"00"）
    #     - "8" vs "B" - 数字8有封闭圆环，字母B有直线
    #     - "5" vs "S" - 数字5有棱角，字母S是弯曲的
    #     - "0" vs "O" - 在十六进制环境中，通常是"0"（零）
    #
    #     验证步骤：
    #     1. 仔细检查任何看起来像重复数字的序列（22, 11, 88）
    #     2. 确认Parameter字段中的管道符"|"不被误认为"1"
    #     3. 在十六进制值中，确保字母A-F被正确识别
    #     4. 寻找上下文线索（TX:, RX:, 十六进制模式如"0x"）
    #     """
    #
    # def enhanced_prompt_template():
    #     """
    #     增强的提示模板，包含更好的OCR指导
    #     """
    #     character_guide = create_character_mapping_prompt()
    #
    #     base_prompt = """你将看到一张图片，图片中可能包含 EOL 测试步骤。如果图片上有以 "TAS" 开头并以五位数字结尾的编码（例如 TAS20678），以及编号旁的灰色字体为Station，
    #     以及编号和灰色字体下的为Program，以及包含如下字段的无框表格：
    #
    #        - Action (操作类型：DoIP_COM、SWITCH、DCV、DCI、RESISTANCE、MATH等)
    #        - LSL
    #        - USL
    #        - Unit
    #        - Circuit
    #        - Parameter(参数：通常包含TX:、RX:格式的十六进制数据)
    #        - Description
    #
    #        请你提取该图片中的所有测试步骤以及可能存在的绿色小箭头标志，；每个步骤提取成一个字典，如果这个步骤有绿色小箭头标志，"Arrow Tag"填充1，如果没有就为空，字段如下：
    #
    #        {{
    #          "Arrow Tag": "...",
    #          "TAS ID": "...",
    #          "Station": "...",
    #          "Program": "...",
    #          "Action": "...",
    #          "LSL": "...",
    #          "USL": "...",
    #          "Unit": "...",
    #          "Circuit": "...",
    #          "Parameter": "...",
    #          "Description": "..."
    #        }}
    #
    #        {character_guide}
    #
    #        额外验证规则：
    #        - Parameter字段通常包含"TX:"和"RX:"后跟十六进制值
    #        - 十六进制值仅使用字符0-9和A-F
    #        - 寻找类似"TX: 22 01 02"和"RX: 62 01 ZZ"的模式（注意：ZZ很常见）
    #        - 竖直条"|"分隔Parameter字段中的TX和RX
    #        - 仔细检查任何可能产生歧义的字符
    #
    #        并将这些字典放入一个列表中返回，最终输出为标准 JSON 格式：
    #
    #        [
    #          {{}},
    #          {{}},
    #          ...
    #        ]
    #
    #        🎯 表格完整性要求：
    #         1. **必须逐行扫描**：从表格第一行到最后一行，不允许遗漏任何数据行
    #         2. **行数完整验证**：仔细数一下表格中的数据行总数，确保提取条目数量完全匹配
    #         3. **特殊行重点关注**：
    #            - 带有绿色小箭头(→)的行特别重要，"Arrow Tag"字段填"1"
    #            - Action为MATH的行容易被忽略，必须提取
    #            - 表格底部的行经常被遗漏，特别注意
    #         4. **边界识别**：每一行都有明确边界，即使行高不一致也要完整识别
    #
    #        ⚠️ 注意事项：
    #        - 如果图片内容不符合上面描述的格式，请直接返回空。
    #        - 一张图片中可能包含多个测试步骤，务必逐行提取**全部**。
    #        - 字段必须按照上面顺序，不可缺失。
    #        - 若某些字段为空，请也保留字段并赋值为空字符串 ""，不要串行。
    #        - 仔细检查字符识别，特别是上述提到的常见错误。
    #        - 返回结果必须是合法完整闭合的 JSON，**不要添加任何解释说明文字或多余注释**。
    #        """
    #
    #     return base_prompt.format(character_guide=character_guide)

    def extract_data_from_image(image_path):
        response = requests.get(image_path, verify=False)
        if response.status_code == 200:
            image_bytes = response.content
            # 🔹 加上预处理
            # processed_bytes = preprocess_image(image_bytes, debug_dir="debug_steps")
            base64_img = base64.b64encode(image_bytes).decode('utf-8')
        else:
            raise Exception(f"Failed to fetch image. Status code: {response.status_code}")
        # with open(image_path, "rb") as f:
        #     image_bytes = f.read()
        #
        # base64_img = base64.b64encode(image_bytes).decode('utf-8')
        # 提取字段 Prompt 模板
        # 使用完善的统一提示
        prompt = create_enhanced_unified_prompt()

        # PROMPT_TEMPLATE = """你将看到一张图片，图片中可能包含 EOL 测试步骤。如果图片上有以 "TAS" 开头并以五位数字结尾的编码（例如 TAS20678），以及编号旁的灰色字体为Station，
        #     以及编号和灰色字体下的为Program，以及包含如下字段的无框表格：
        #
        #        - Action
        #        - LSL
        #        - USL
        #        - Unit
        #        - Circuit
        #        - Parameter
        #        - Description
        #
        #        请你提取该图片中的所有测试步骤以及可能存在的绿色小箭头标志，；每个步骤提取成一个字典，如果这个步骤有绿色小箭头标志，“Arrow Tag"填充1，如果没有就为空，字段如下：
        #
        #        {
        #          "Arrow Tag": "...",
        #          "TAS ID": "...",
        #          "Station": "...",
        #          "Program": "...",
        #          "Action": "...",
        #          "LSL": "...",
        #          "USL": "...",
        #          "Unit": "...",
        #          "Circuit": "...",
        #          "Parameter": "...",
        #          "Description": "..."
        #        }
        #
        #        并将这些字典放入一个列表中返回，最终输出为标准 JSON 格式：
        #
        #        [
        #          {...},
        #          {...},
        #          ...
        #        ]
        #
        #        ⚠️ 注意事项：
        #        - 如果图片内容不符合上面描述的格式，请直接返回空。
        #        - 一张图片中可能包含多个测试步骤，务必提取**全部**。
        #        - 字段必须按照上面顺序，不可缺失。
        #        - 若某些字段为空，请也保留字段并赋值为空字符串 ""，不要串行。
        #        - 请不要误识别，例如zz不是22，|不是1，6E不是8E，0D不是00等。
        #        - 返回结果必须是合法完整闭合的 JSON，**不要添加任何解释说明文字或多余注释**。
        #        """

        response = azure_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": """你是高级技术文档解析专家，专精于：

核心专长：
1. 完整表格数据提取 - 绝不遗漏任何行，确保100%完整性
2. 精准字符识别 - 特别擅长区分zz vs 22、| vs 1等易混淆字符
3. 十六进制数据理解 - 熟悉TX/RX格式和技术参数结构
4. 视觉标记识别 - 准确识别绿色箭头等特殊标记

工作标准：
- 完整性第一：宁可重复确认也不能遗漏数据
- 准确性关键：每个字符都关系到系统正常运行
- 系统性验证：严格按照检查清单逐项核实

请按照提示要求，以最高标准完成技术文档解析任务。"""
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{base64_img}"
                            }
                        }
                    ]
                }
            ],
            temperature=0,
            top_p=0.1  # 进一步提高确定性
        )
        raw_content = response.choices[0].message.content
        return raw_content

        # 获取当前文件的绝对路径

    current_file_path = os.path.abspath(__file__)
    # 获取当前文件的根目录
    root_directory = os.path.dirname(current_file_path)
    # 收集所有出现的station 分station存储测试结果
    station_collection_list = []
    records = {}
    doc = fitz.open(pdf_path)
    for i in range(len(doc)):
        page = doc[i]
        pix = page.get_pixmap(dpi=200)
        image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{i + 1}.png"
        # docker 存储 image
        local_img_path = os.path.join(root_directory, 'prv_pdf_images', f"{image_filename}")
        print("local_img_path: ", local_img_path)
        pix.save(local_img_path)
        # with open(local_img_path, 'wb') as file:
        #     file.write(pix)
        # 文件服务器存储image
        with open(local_img_path, "rb") as image:
            files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
            upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
            response = requests.post(upload_url, files=files, verify=False)
            image_path = response.json()['data'][0]['url']

            try:
                # 开始获取每页prv的内容
                json_str = extract_data_from_image(image_path)
                print("🔍 LLM 原始返回内容：", repr(json_str))

                json_str_clean = clean_llm_json(json_str)
                print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容

                if json_str_clean.strip():  # 防止空字符串
                    data_list = json.loads(json_str_clean)
                    print("返回的prv表格: ", data_list)

                    page_testing_matrix = []

                    # 收集本页所有的 tas id 用来 by tas number 分组
                    tas_id_list = []
                    for data in data_list:
                        tas_id_list.append(data['TAS ID'])
                    tas_id_unique_list = list(set(tas_id_list))

                    # ***** 将同一个tas number分成一组 成为一个in_use_data_list 然后进行规则匹配 *****
                    for tas_id in tas_id_unique_list:
                        in_use_data_list = []
                        for data_list_row in data_list:
                            if data_list_row['TAS ID'] == tas_id:
                                in_use_data_list.append(data_list_row)
                        print('TES ID: ', tas_id, 'Station: ', in_use_data_list[0]['Station'], 'RELATED DATA LIST: ', in_use_data_list)

                        testing_matrix_dict = {}
                        group_matrix_list = []
                        # 开始循环遍历一个tas number 下的所有action项(in_use_data_list) 进行规则匹配 生成testing
                        for in_use_row_index, in_use_row in enumerate(in_use_data_list):

                            single_matrix_list = []
                            last_row = {}
                            parameter = in_use_row.get('Parameter', '')
                            parts = parameter.split('|')
                            std_tx = parts[0].lower() if len(parts) > 0 else ""
                            std_rx = parts[1].lower() if len(parts) > 1 else ""

                            std_description = in_use_row.get('Description', '')
                            std_description = std_description.replace(' ', '').lower()

                            std_circuit = in_use_row.get('Circuit', '').lower()
                            print(f"std_tx: {std_tx}, std_rx: {std_rx}, std_description: {std_description}")

                            # 【DoIP_COM】
                            if in_use_row.get('Action').upper() == 'DOIP_COM':
                                # 场景一： 【DoIP_COM】 set byte 判断 std_tx中有xx 并且 std_description中有setbyte (注意处理大小写空格问题)
                                if 'xx' in std_tx and 'setbyte' in std_description:
                                    print("匹配场景一")
                                    # 拆分 parameter
                                    parts = in_use_row.get('Parameter', '').split('|')

                                    # ---- RX ----
                                    if len(parts) > 1:
                                        replace_rx = parts[1].upper().replace('RX:', '').strip()
                                        print("场景一：找到rx的内容")
                                    else:
                                        replace_rx = ""
                                        print("场景一：没有找到rx的内容")

                                    # ---- TX ----
                                    pattern = re.compile(r'0x\s*([0-9a-f]{2})', re.IGNORECASE)
                                    replace_xx = pattern.search(in_use_row.get('Description', ''))

                                    if replace_xx:
                                        replace_value = replace_xx.group(1).upper()
                                        print(f"场景一：setbyte中匹配到description有需要的2位数为 {replace_value}")
                                    else:
                                        replace_value = "00"
                                        print("场景一：setbyte中没有匹配到description有需要的2位数，暂时用00替代")

                                    tx_part = parts[0] if parts else ""
                                    replace_tx = tx_part.upper().replace('TX:', '').replace('XX', replace_value).strip()

                                    # ---- RX 长度 ----
                                    replace_rx_length = str(len(replace_rx.replace(' ', '')) // 2)

                                    single_matrix_list = [{"Test Id": "",
                                                     "Input Message": "sDoIP_SND_RCV_RC",
                                                     "Description": "",
                                                     "Bitmap": "",
                                                     "speImg_xy": "",
                                                     "LoLimit": replace_rx,
                                                     "HiLimit": replace_rx,
                                                     "Unit": "String",
                                                     "speJumpPass": "",
                                                     "speJumpFail": "",
                                                     "speFailGo": "",
                                                     "speReference": "",
                                                     "speSend": replace_tx,
                                                     "speReceive": '"%z' + replace_rx_length + '"',
                                                     "speBusTimeOut": "",
                                                     "speRepeat": "10",
                                                     "speTimeOut": "",
                                                     "speDelay": "",
                                                     "speDelayRepeat": "",
                                                     "speResult": "",
                                                     "speCycleCounter": "",
                                                     "speRegisterTransfer": "",
                                                     "speParamString": "",
                                                     "speResultString": "",
                                                     "speBlockCode": "",
                                                     "PromptMessage": "",
                                                     "LastColumn": ""}]
                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[0]["Test Id"] = tas_id



                                # 场景二： 【DoIP_COM】 check byte ** 有需要出 test id再补充的地方 **
                                elif 'zz' in std_rx and 'checkbyte' in std_description:
                                    print("匹配场景二")
                                    # 拆分 parameter
                                    parts = in_use_row.get('Parameter', '').split('|')

                                    # ---- row1 ----
                                    # 填充 rx 到 lolimit & hilimit  填充 tx 到 spesend
                                    rx_part = parts[1] if len(parts) > 1 else ""
                                    replace_rx = rx_part.upper().replace('RX:', '').replace(' ZZ', '').strip()
                                    replace_tx = (parts[0] if parts else "").upper().replace('TX:', '').strip()
                                    # 填充 rx 的长度到 spereceive (整除)
                                    replace_rx_length = str(len(replace_rx.replace(' ', '')) // 2)

                                    # ---- row3 ----
                                    replace_lsl = in_use_row.get('LSL', '')
                                    replace_usl = in_use_row.get('USL', '')
                                    replace_unit = in_use_row.get('Unit', '')

                                    # ---- spe reference ----
                                    # 填充check byte 后的数字到 spe reference
                                    # 考虑两种情况 一种是有一个数字区间 一种是一个数字
                                    pattern = re.compile(r'check\s*byte\s*(\d+)(?:\s*[-~]\s*(\d+))?', re.IGNORECASE)
                                    match = pattern.search(in_use_row.get('Description', ''))

                                    if match:
                                        byte_start = int(match.group(1))
                                        byte_end = int(match.group(2)) if match.group(2) else byte_start

                                        byte_length = byte_end - byte_start + 1
                                        first_number = byte_start * 3
                                        second_number = byte_length * 3 - 1

                                        print(f"Byte range: {byte_start} to {byte_end}, length: {byte_length}")
                                        replace_description_number = f'"%s*{first_number}%s{second_number}"'
                                    else:
                                        print("No match found.")
                                        replace_description_number = ""
                                    # speresult需要填充上一行的test id 需要在一组tas number的action都遍历完成之后处理
                                    replace_tag = '2_check_byte_tag'

                                    single_matrix_list = [{"Test Id": "",
                                                     "Input Message": "sDoIP_SND_RCV_RC",
                                                     "Description": "",
                                                     "Bitmap": "",
                                                     "speImg_xy": "",
                                                     "LoLimit": replace_rx,
                                                     "HiLimit": replace_rx,
                                                     "Unit": "String",
                                                     "speJumpPass": "",
                                                     "speJumpFail": "",
                                                     "speFailGo": "",
                                                     "speReference": "",
                                                     "speSend": replace_tx,
                                                     "speReceive": '"%z' + replace_rx_length + '"',
                                                     "speBusTimeOut": "",
                                                     "speRepeat": "10",
                                                     "speTimeOut": "",
                                                     "speDelay": "",
                                                     "speDelayRepeat": "",
                                                     "speResult": "",
                                                     "speCycleCounter": "",
                                                     "speRegisterTransfer": "",
                                                     "speParamString": "",
                                                     "speResultString": "",
                                                     "speBlockCode": "",
                                                     "PromptMessage": "",
                                                     "LastColumn": ""},
                                                          {"Test Id": "",
                                                       "Input Message": "sUTIL_VerifyResult",
                                                       "Description": "",
                                                       "Bitmap": "",
                                                       "speImg_xy": "",
                                                       "LoLimit": "%s*",
                                                       "HiLimit": "%s*",
                                                       "Unit": "String",
                                                       "speJumpPass": "",
                                                       "speJumpFail": "",
                                                       "speFailGo": "",
                                                       "speReference": "",
                                                       "speSend": "",
                                                       "speReceive": "",
                                                       "speBusTimeOut": "",
                                                       "speRepeat": "",
                                                       "speTimeOut": "",
                                                       "speDelay": "",
                                                       "speDelayRepeat": "",
                                                       "speResult": "",
                                                       "speCycleCounter": "",
                                                       "speRegisterTransfer": "Rm_DoIP_Param1",
                                                       "speParamString": "RESULT_TYPE\n{String}",
                                                       "speResultString": "",
                                                       "speBlockCode": "",
                                                       "PromptMessage": "",
                                                       "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_ParseTASResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_lsl,
                                                           "HiLimit": replace_usl,
                                                           "Unit": replace_unit,
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": replace_description_number,
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": replace_tag,
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "RESULT_TYPE\n{Hex}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]

                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[2]["Test Id"] = tas_id


                                # 场景三： 【DoIP_COM】 Compare with  SOS file
                                elif 'zz' in std_rx and 'comparewithsosfile' in std_description:
                                    print("匹配场景三")
                                    # ---- row1 ----
                                    pattern = re.compile(r'compare[\s\-_]*with[\s\-_]*sos[\s\-_]*file\s*[:：]\s*(.*)', re.IGNORECASE)
                                    match = pattern.search(in_use_row.get('Description', ''))
                                    replace_description = match.group(1).strip() if match else ""
                                    print(f'replace_description: {replace_description}')

                                    # ---- row3 ----
                                    parts = in_use_row.get('Parameter', '').split('|')
                                    rx_part = parts[1] if len(parts) > 1 else ""
                                    tx_part = parts[0] if parts else ""
                                    replace_rx = rx_part.upper().replace('RX:', '').replace(' ZZ', '').strip()
                                    replace_tx = tx_part.upper().replace('TX:', '').strip()

                                    replace_rx_length = str(len(replace_rx.replace(' ', '')) // 2)

                                    # ---- row5 ----
                                    replace_lsl = in_use_row.get('LSL', '')
                                    replace_usl = in_use_row.get('USL', '')
                                    replace_unit = in_use_row.get('Unit', '')

                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": "sUTIL_GetVersionData",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "0",
                                                           "HiLimit": "0",
                                                           "Unit": "Dec",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rg_Value",
                                                           "speParamString": "VERSION_KEY{" + replace_description +"}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_VerifyResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "%s*",
                                                           "HiLimit": "%s*",
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rg_Value",
                                                           "speParamString": "RESULT_TYPE\n{String}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sDoIP_SND_RCV_RC",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_rx,
                                                           "HiLimit": replace_rx,
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": replace_tx,
                                                           "speReceive": '"%z' + replace_rx_length + '"',
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "10",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_VerifyResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "%s*",
                                                           "HiLimit": "%s*",
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rm_DoIP_Param1",
                                                           "speParamString": "RESULT_TYPE\n{String}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_Compare_Strings_0",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_lsl,
                                                           "HiLimit": replace_usl,
                                                           "Unit": replace_unit,
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "STRING1 {@Rg_Value@} STRING2 {@Rm_DoIP_Param1@} COMPARE_STRING {String1_String2} TO_UPPER {0}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]

                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[4]["Test Id"] = tas_id



                                # 场景四： 【DoIP_COM】 Compare with MIS
                                elif 'zz' in std_rx and 'comparewithmis' in std_description:
                                    print("匹配场景四")
                                    # ---- row1 ----
                                    pattern = re.compile(r'compare[\s\-_]*with[\s\-_]*sos[\s\-_]*file\s*[:：]\s*(.*)', re.IGNORECASE)
                                    match = pattern.search(in_use_row.get('Description', ''))
                                    replace_description = match.group(1).strip() if match else ""
                                    print(f'replace_description: {replace_description}')

                                    # ---- row3 ----
                                    parts = in_use_row.get('Parameter', '').split('|')
                                    rx_part = parts[1] if len(parts) > 1 else ""
                                    tx_part = parts[0] if parts else ""
                                    replace_rx = rx_part.upper().replace('RX:', '').replace(' ZZ', '').strip()
                                    replace_tx = tx_part.upper().replace('TX:', '').strip()

                                    replace_rx_length = str(len(replace_rx.replace(' ', '')) // 2)

                                    # ---- row5 ----
                                    replace_lsl = in_use_row.get('LSL', '')
                                    replace_usl = in_use_row.get('USL', '')
                                    replace_unit = in_use_row.get('Unit', '')

                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": "sUTIL_ReadParameterKey",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "%s*",
                                                           "HiLimit": "%s*",
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rg_Value",
                                                           "speParamString": "VERSION_KEY{" + replace_description + "}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_VerifyResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "%s*",
                                                           "HiLimit": "%s*",
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rg_Value",
                                                           "speParamString": "RESULT_TYPE\n{String}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sDoIP_SND_RCV_RC",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_rx,
                                                           "HiLimit": replace_rx,
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": replace_tx,
                                                           "speReceive": '"%z' + replace_rx_length + '"',
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "10",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_VerifyResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "%s*",
                                                           "HiLimit": "%s*",
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rm_DoIP_Param1",
                                                           "speParamString": "RESULT_TYPE\n{String}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_Compare_Strings_0",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_lsl,
                                                           "HiLimit": replace_usl,
                                                           "Unit": replace_unit,
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "STRING1 {@Rg_Value@} STRING2 {@Rm_DoIP_Param1@} COMPARE_STRING {String1_String2} TO_UPPER {0}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]

                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[4]["Test Id"] = tas_id



                                #场景五： 【DoIP_COM】 rx中包含zz 但是description不包含 'checkbyte' & 'comparewithsosfile' & 'comparewithmis'
                                elif 'zz' in std_rx and all(x not in std_description for x in['checkbyte', 'comparewithsosfile', 'comparewithmis']):
                                    print("匹配场景五")
                                    # ---- row1 ----
                                    parts = in_use_row.get('Parameter', '').split('|')
                                    rx_part = parts[1] if len(parts) > 1 else ""
                                    tx_part = parts[0] if parts else ""

                                    replace_rx = rx_part.upper().replace('RX:', '').replace(' ZZ', '').strip()
                                    replace_tx = tx_part.upper().replace('TX:', '').strip()

                                    replace_rx_length = str(len(replace_rx.replace(' ', '')) // 2)

                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": "sDoIP_SND_RCV_RC",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_rx,
                                                           "HiLimit": replace_rx,
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": replace_tx,
                                                           "speReceive": '"%z' + replace_rx_length + '"',
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "10",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sUTIL_VerifyResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "%s*",
                                                           "HiLimit": "%s*",
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "Rm_DoIP_Param1",
                                                           "speParamString": "RESULT_TYPE\n{String}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]

                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[1]["Test Id"] = tas_id

                                #场景六： 【DoIP_COM】 tx和rx中没有xx和zz 同时 description中没有出现 'setbyte' & 'checkbyte' & 'comparewithsosfile' & 'comparewithmis'
                                elif 'xx' not in std_tx and 'zz' not in std_rx and all(x not in std_description for x in['setbyte','checkbyte', 'comparewithsosfile', 'comparewithmis']):
                                    print("匹配场景六")
                                    # ---- row1 ----
                                    parts = in_use_row.get('Parameter', '').split('|')
                                    rx_part = parts[1] if len(parts) > 1 else ""
                                    tx_part = parts[0] if parts else ""

                                    replace_rx = rx_part.upper().replace('RX:', '').strip()
                                    replace_tx = tx_part.upper().replace('TX:', '').strip()

                                    replace_rx_length = str(len(replace_rx.replace(' ', '')) // 2)
                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": "sDoIP_SND_RCV_RC",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_rx,
                                                           "HiLimit": replace_rx,
                                                           "Unit": "String",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": replace_tx,
                                                           "speReceive": '"%z' + replace_rx_length + '"',
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "10",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]
                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[0]["Test Id"] = tas_id

                            # 【SWITCH】
                            elif in_use_row.get('Action').upper() == 'SWITCH':
                                print("匹配SWITCH")
                                # 场景七：currentsink
                                if 'currentsink' in std_tx:
                                    print("匹配场景七")
                                    circuit_value = in_use_row.get('Circuit', '')
                                    if circuit_value:
                                        row = excel_table.loc[
                                            excel_table['Circuit'].str.lower() == circuit_value.lower()]
                                    else:
                                        row = pd.DataFrame()

                                    if not row.empty:
                                        replace_inputmessage = row.iloc[0].get('Input Message', '')
                                        replace_inputmessage_for_off = replace_inputmessage.replace('Set', 'Clear')
                                        replace_speparamstring = row.iloc[0].get('speParamString', '')
                                    else:
                                        replace_inputmessage = ''
                                        replace_speparamstring = ''

                                    if 'off' in std_tx:
                                        single_matrix_list = [{"Test Id": "",
                                                               "Input Message": "sLoad_Chroma63610_LoadOFF",
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "0.00",
                                                               "HiLimit": "0.00",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": "",
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""},
                                                              {"Test Id": "",
                                                               "Input Message": "sREL_FITS_TokenReturn",
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "0.00",
                                                               "HiLimit": "0.00",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": "FITSRequestEquipment\n{ELOAD}",
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""},
                                                              {"Test Id": "",
                                                               "Input Message": replace_inputmessage,
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "1",
                                                               "HiLimit": "1",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": replace_speparamstring,
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""}
                                                              ]
                                    else:
                                        # row3: 替换 sink{}里的数字
                                        pattern = re.compile(r"=\s*(\d+(?:\.\d+)?)\s*(mA|A)", re.IGNORECASE)
                                        param_text = in_use_row.get('Parameter', '')
                                        match = pattern.search(param_text)

                                        if match:
                                            number = float(match.group(1))
                                            unit = match.group(2).strip().upper()
                                            if unit == 'A':
                                                replace_unit = str(number)
                                            elif unit == 'MA':
                                                replace_unit = str(number / 1000)
                                            else:
                                                replace_unit = ''
                                        else:
                                            replace_unit = ''

                                        single_matrix_list = [{"Test Id": "",
                                                               "Input Message": replace_inputmessage,
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "1",
                                                               "HiLimit": "1",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": replace_speparamstring,
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""},
                                                              {"Test Id": "",
                                                               "Input Message": "sREL_FITS_TokenRequest",
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "0.00",
                                                               "HiLimit": "0.00",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": "FITSRequestEquipment\n{ELOAD}",
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""},
                                                              {"Test Id": "",
                                                               "Input Message": "sLoad_Chroma63610_CurrentSink",
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "0.00",
                                                               "HiLimit": "0.00",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": "SINK{" + replace_unit + "}",
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""},
                                                              {"Test Id": "",
                                                               "Input Message": "sLoad_Chroma63610_LoadON",
                                                               "Description": "",
                                                               "Bitmap": "",
                                                               "speImg_xy": "",
                                                               "LoLimit": "0.00",
                                                               "HiLimit": "0.00",
                                                               "Unit": "Dec",
                                                               "speJumpPass": "",
                                                               "speJumpFail": "",
                                                               "speFailGo": "",
                                                               "speReference": "",
                                                               "speSend": "",
                                                               "speReceive": "",
                                                               "speBusTimeOut": "",
                                                               "speRepeat": "",
                                                               "speTimeOut": "",
                                                               "speDelay": "",
                                                               "speDelayRepeat": "",
                                                               "speResult": "",
                                                               "speCycleCounter": "",
                                                               "speRegisterTransfer": "",
                                                               "speParamString": "MODE\n{MODE CCM}",
                                                               "speResultString": "",
                                                               "speBlockCode": "",
                                                               "PromptMessage": "",
                                                               "LastColumn": ""}
                                                              ]

                                # 场景八：relay set  ** 有需要出一组testing齐了再补充的地方 **
                                elif 'relayset' in std_description:
                                    print("匹配场景八")
                                    # 识别 circuit 和 description 去 additional file 获取 input message & speparamstring
                                    replace_circuit = in_use_row.get('Circuit', '')
                                    replace_description = in_use_row.get('Description', '')
                                    replace_circuit_description = (replace_circuit + ' ' +replace_description).lower()
                                    print(f'replace_circuit_description: {replace_circuit_description}')

                                    row = excel_table.loc[excel_table['Circuit'].str.lower() == replace_circuit_description]
                                    print(f"row:{row}")

                                    if not row.empty:
                                        replace_inputmessage = row.iloc[0].get('Input Message', "")
                                        replace_speparamstring = row.iloc[0].get('speParamString', "")
                                        print(f"row not empty can find excel info ,replace_inputmessage: {replace_inputmessage}, replace_speparamstring:{replace_speparamstring}")
                                    else:
                                        replace_inputmessage = ""
                                        replace_speparamstring = ""
                                        print(f"row empty")

                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": replace_inputmessage,
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "1",
                                                           "HiLimit": "1",
                                                           "Unit": "Dec",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": replace_speparamstring,
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}]
                                    # 最后一行的inputmessage需要把bitset 改成 bitclear
                                    replace_inputmessage_bitclear = replace_inputmessage.replace('Set', 'Clear')
                                    print(f"replace_inputmessage_bitclear: {replace_inputmessage_bitclear}")
                                    last_row = {"Test Id": "",
                                               "Input Message": replace_inputmessage_bitclear,
                                               "Description": "",
                                               "Bitmap": "",
                                               "speImg_xy": "",
                                               "LoLimit": "1",
                                               "HiLimit": "1",
                                               "Unit": "Dec",
                                               "speJumpPass": "",
                                               "speJumpFail": "",
                                               "speFailGo": "",
                                               "speReference": "",
                                               "speSend": "",
                                               "speReceive": "",
                                               "speBusTimeOut": "",
                                               "speRepeat": "",
                                               "speTimeOut": "",
                                               "speDelay": "",
                                               "speDelayRepeat": "",
                                               "speResult": "",
                                               "speCycleCounter": "",
                                               "speRegisterTransfer": "",
                                               "speParamString": replace_speparamstring,
                                               "speResultString": "",
                                               "speBlockCode": "",
                                               "PromptMessage": "",
                                               "LastColumn": ""}

                            # 【DCV/DCI/RESISTANCE】
                            action = in_use_row.get('Action', '').upper()
                            if any(x in action for x in ['DCV', 'DCI', 'RESISTANCE']):
                                # 场景九：measure
                                if 'measure' in std_description:
                                    print("匹配场景九")
                                    replace_circuit = in_use_row.get('Circuit', '')
                                    row = excel_table.loc[excel_table[
                                                              'Circuit'].str.upper() == replace_circuit.upper()] if replace_circuit else pd.DataFrame()

                                    if not row.empty:
                                        replace_inputmessage = row.iloc[0].get('Input Message', '')
                                        replace_speregistertransfer = row.iloc[0].get('speRegisterTransfer', '')
                                        replace_speparamstring = row.iloc[0].get('speParamString', '')
                                        replace_inputmessage_bitclear = replace_inputmessage.replace('Set',
                                                                                                     'Clear') if replace_inputmessage else ''
                                    else:
                                        replace_inputmessage = ''
                                        replace_speregistertransfer = ''
                                        replace_speparamstring = ''
                                        replace_inputmessage_bitclear = ''

                                    # row3
                                    replace_lsl = in_use_row.get('LSL', '')
                                    replace_usl = in_use_row.get('USL', '')
                                    replace_unit = in_use_row.get('Unit', '')

                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": replace_inputmessage,
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "1",
                                                           "HiLimit": "1",
                                                           "Unit": "Dec",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": replace_speparamstring,
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sREL_FITS_TokenRequest",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "0.00",
                                                           "HiLimit": "0.00",
                                                           "Unit": "Dec",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "FITSRequestEquipment\n{DMM}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "mSpe_DMM1_MeasureLoop",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_lsl,
                                                           "HiLimit": replace_usl,
                                                           "Unit": replace_unit,
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": replace_speregistertransfer,
                                                           "speParamString": "Rg_DMM_Tx",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": "sREL_FITS_TokenReturn",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "0.00",
                                                           "HiLimit": "0.00",
                                                           "Unit": "Dec",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "FITSRequestEquipment\n{DMM}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""},
                                                          {"Test Id": "",
                                                           "Input Message": replace_inputmessage_bitclear,
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": "1",
                                                           "HiLimit": "1",
                                                           "Unit": "Dec",
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": "",
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": replace_speparamstring,
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]

                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[2]["Test Id"] = tas_id



                            # 【MATH】
                            elif in_use_row.get('Action', '').upper() == 'MATH':
                                # 场景十：checkbyte ** 有最后处理的**
                                if 'checkbyte' in std_description.lower():
                                    print("匹配场景十")
                                    replace_lsl = in_use_row.get('LSL', '')
                                    replace_usl = in_use_row.get('USL', '')
                                    replace_unit = in_use_row.get('Unit', '')

                                    # 填充 check byte 数字
                                    pattern = re.compile(r'check\s*byte\s*(\d+)(?:\s*[-~]\s*(\d+))?', re.IGNORECASE)
                                    desc_text = in_use_row.get('Description', '')
                                    match = pattern.search(desc_text)

                                    if match:
                                        byte_start = int(match.group(1))
                                        byte_end = int(match.group(2)) if match.group(2) else byte_start
                                        byte_length = byte_end - byte_start + 1

                                        first_number = byte_start * 3
                                        second_number = (byte_length * 3 - 1) if match.group(2) else 2

                                        replace_description_number = f'"%s*{first_number}%s{second_number}"'
                                        print(f"Byte range: {byte_start} to {byte_end}, length: {byte_length}")
                                    else:
                                        print("No match found.")
                                        replace_description_number = ""

                                    single_matrix_list = [{"Test Id": "",
                                                           "Input Message": "sUTIL_ParseTASResult",
                                                           "Description": "",
                                                           "Bitmap": "",
                                                           "speImg_xy": "",
                                                           "LoLimit": replace_lsl,
                                                           "HiLimit": replace_usl,
                                                           "Unit": replace_unit,
                                                           "speJumpPass": "",
                                                           "speJumpFail": "",
                                                           "speFailGo": "",
                                                           "speReference": replace_description_number,
                                                           "speSend": "",
                                                           "speReceive": "",
                                                           "speBusTimeOut": "",
                                                           "speRepeat": "",
                                                           "speTimeOut": "",
                                                           "speDelay": "",
                                                           "speDelayRepeat": "",
                                                           "speResult": "MATHTAG",
                                                           "speCycleCounter": "",
                                                           "speRegisterTransfer": "",
                                                           "speParamString": "RESULT_TYPE\n{Hex}",
                                                           "speResultString": "",
                                                           "speBlockCode": "",
                                                           "PromptMessage": "",
                                                           "LastColumn": ""}
                                                          ]

                                    # 通过箭头确认 tas number 的填充
                                    if in_use_row['Arrow Tag'] == "1":
                                        print("find arrow tag in this row, need give the tas number")
                                        single_matrix_list[0]["Test Id"] = tas_id


                            group_matrix_list.extend(single_matrix_list)

                        if group_matrix_list:
                            # 判断是否有情况8的last row， 如果有就需要在本组结束后加上
                            if last_row:
                                print("有last row需要加")
                                group_matrix_list.append(last_row)

                            # 填充一组的第一列 tas number 还需要判断是否为空 因为主tasnumber已经填充过了
                            tas_number_suffix = 0
                            for row_in_group_matrix_list in group_matrix_list:
                                test_id_value = row_in_group_matrix_list.get("Test Id")
                                if test_id_value != tas_id:
                                    tas_number_suffix += 1
                                    print(f"开始插入tas_id:{tas_id}.{tas_number_suffix:02d}")
                                    row_in_group_matrix_list["Test Id"] = f'{tas_id}.{tas_number_suffix:02d}'

                            # 对于 speResult 有 MATHTAG 需要替换成DoIP sUTIL_VerifyResult的 number
                            def find_first_index(df: pd.DataFrame, column: str, value) -> Optional[int]:
                                """返回 DataFrame 中某列第一次出现特定内容的行索引，如果没找到返回 None"""
                                mask = df[column].eq(value)
                                if mask.any():
                                    return mask.idxmax()
                                return None

                            # 找到 group_matrix_list 中第一个 MATHTAG 对应的 Test Id
                            first_row_index_of_math_tag = find_first_index(pd.DataFrame(group_matrix_list), "speResult",
                                                                           "MATHTAG")
                            change_tas_number_of_math_tag = None
                            if first_row_index_of_math_tag is not None and first_row_index_of_math_tag > 0:
                                # 取前一行的 "Test Id"
                                change_tas_number_of_math_tag = group_matrix_list[first_row_index_of_math_tag-1].get(
                                    "Test Id")

                            # 替换 group_matrix_list 中的 MATHTAG
                            if change_tas_number_of_math_tag:
                                for row in group_matrix_list:
                                    if row.get("speResult") == "MATHTAG":
                                        row["speResult"] = change_tas_number_of_math_tag

                            # 找到 group_matrix_list 中 speResult 列 是否有 2_check_byte_tag
                            first_row_index_of_2_check_byte_tag = find_first_index(pd.DataFrame(group_matrix_list), "speResult",
                                                                           "2_check_byte_tag")
                            change_tas_number_of_2_check_byte_tag = None
                            if first_row_index_of_2_check_byte_tag is not None:
                                change_tas_number_of_2_check_byte_tag = group_matrix_list[first_row_index_of_2_check_byte_tag - 1].get(
                                    "Test Id")
                            #替换 group_matrix_list 中的 2_check_byte_tag
                            if first_row_index_of_2_check_byte_tag:
                                for row in group_matrix_list:
                                    if row.get("speResult") == "2_check_byte_tag":
                                        row["speResult"] = change_tas_number_of_2_check_byte_tag
                            print(f"tas_number {tas_id},返回的testing表格: {group_matrix_list}")

                        # 存入字典
                        testing_matrix_dict['station'] = in_use_row['Station']
                        testing_matrix_dict['category'] = in_use_row['Program']
                        testing_matrix_dict['testing_content'] = group_matrix_list
                        print(f"tas_number {tas_id},返回的testing表格: {testing_matrix_dict}")

                        page_testing_matrix.append(testing_matrix_dict)


                    # 异步传输每个image和content
                    result_dict = {'page_image_path': image_path, 'page_image_content': data_list,
                                   'testing_matrix': page_testing_matrix}
                    print(f"！！！ 发送给前端的结果：{result_dict}")
                    prv_message = {'connectionID': task_id,
                                   'category': 'prv_image', 'from': '', 'to': '',
                                   'message': json.dumps(result_dict, ensure_ascii=False),
                                   'remarks': json.dumps({'paragraph_start': 1,
                                                          'response_end': 0})}
                    try:
                        asyncio.run(websocket_client(prv_message))
                    except:
                        time.sleep(0.5)
                        asyncio.run(websocket_client(prv_message))

                    def extend_with_label(station_records, label, new_data):
                        """
                        在 station_records 中插入一个标签行，并扩展新的测试内容
                        """
                        if new_data:
                            # 插入一条标识占行
                            station_records.append({'Test Id': label})
                            # 插入真实测试内容
                            station_records.extend(new_data)
                        return station_records

                    # 遍历每一页的测试矩阵
                    if page_testing_matrix:
                        for testing_group in page_testing_matrix:
                            station = testing_group['station']
                            label = testing_group['category']
                            actual_testing_content = testing_group['testing_content']

                            # 初始化 station 的记录列表（如果不存在）
                            if station not in records:
                                records[station] = []

                            # 添加占行信息 + 实际测试内容
                            records[station] = extend_with_label(records[station], label, actual_testing_content)


                else:
                    print(f"⚠️ 空内容跳过解析：{image_path}")

            except Exception as e:
                print(f"❌ Error parsing {image_path}: {e}")
                # print(f"⚠️ 原始返回内容为： {repr(json_str)}")



    # 结果生成excel存储到github
    for station, testing_content in records.items():
        print(f"content save into excel, station: {station}, testing_content: {testing_content}")
        if testing_content:
            print(f"{station} has content to write in excel and download!!!")

            df = pd.DataFrame(testing_content)
            # all_keys = set()
            # for r in testing_content:
            #     all_keys.update(r.keys())
            # print(f"all_keys{all_keys}")
            #
            # # 字段统一化处理
            # normalized_testing_content = []
            # for r in testing_content:
            #     normalized_testing_content.append({key: r.get(key, None) for key in all_keys})
            # print(f"normalized_testing_content{normalized_testing_content}")
            # df = pd.DataFrame(normalized_testing_content)

            # 生成 Excel 文件名（包含 station）
            excel_name = f"{station}_{os.path.basename(pdf_path).replace('.pdf', '.xlsx')}"
            print("download excel name:", excel_name)

            # 确保输出目录存在
            output_dir = os.path.join(root_directory, 'prv_pdf_to_testing_excel')
            os.makedirs(output_dir, exist_ok=True)

            # 拼接完整路径
            output_path = os.path.join(output_dir, excel_name)

            # 保存 DataFrame
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="spe")
                ws = writer.sheets["spe"]
                # 在这里直接对 ws 操作
                max_col = ws.max_column

                # 获取 "Test Id" 所在列（容错）
                header = [cell.value for cell in ws[1]]
                try:
                    test_id_col_idx = header.index("Test Id") + 1 # openpyxl 列号从 1 开始
                except StopIteration:
                    raise Exception("❌ 没有找到列名 'Test Id'，请确认表头是否正确")


                # 遍历每一行（从第2行开始）
                for row in range(2, ws.max_row + 1):
                    test_id_value = ws.cell(row=row, column=test_id_col_idx).value
                    # 除 Test Id 外其它列是否全空？
                    others_empty = all(
                        ws.cell(row=row, column=col).value in [None, ""]
                        for col in range(1, max_col + 1) if col != test_id_col_idx
                    )

                    if test_id_value and others_empty:
                        # 合并整行（从A列到最后一列）
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
                        # 设置居中
                        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=row, column=1).font = Font(bold=True)

                # # ✅ 统一给所有单元格加上 wrapText
                # for row in ws.iter_rows():
                #     for cell in row:
                #         cell.alignment = Alignment(
                #             horizontal=cell.alignment.horizontal,
                #             vertical=cell.alignment.vertical,
                #             wrapText=True
                #         )

            # 保存
            download_excel_path_dict[excel_name] = output_path
            print("✅ Excel 已保存到:", output_path)

    # 异步传输downloadexcel地址信息
    download_message = {'connectionID': task_id,
                   'category': 'downloadfile', 'from': '', 'to': '',
                   'message': json.dumps(download_excel_path_dict, ensure_ascii=False),
                    'remarks': json.dumps({'paragraph_start': 1,
                                        'response_end': 0})}
    try:
        asyncio.run(websocket_client(download_message))
    except:
        time.sleep(0.5)
        asyncio.run(websocket_client(download_message))

    # 异步传输结束信号空字符串
    prv_message = {'connectionID': task_id,
                   'category': 'text', 'from': '', 'to': '',
                   'message': '',
                   'remarks': json.dumps({'paragraph_start': 0,
                                          'response_end': 1})}
    try:
        asyncio.run(websocket_client(prv_message))
    except:
        time.sleep(0.5)
        asyncio.run(websocket_client(prv_message))

#######################################################################################################

                              #  TCD ELECTRIC ARCHITECTURE   #

#######################################################################################################

def find_electric_architecture_page(pdf_path):
    section_title = "3.2.2 ELECTRICAL INTERFACE"
    count = 0
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text and section_title.lower() in text.lower():
                count += 1
                # 需要考虑目录会出现 所以取第二个出现的位置
                if count == 2:
                    return i
    return None


def auto_tcd_electric_architecture_extract(task_id, pdf_path, start_page):

    def clean_llm_json(json_str):
        # 去掉前后的 markdown 代码块标识符 ```json ... ```
        if json_str.startswith("```"):
            json_str = re.sub(r"^```(?:json)?\n", "", json_str)
            json_str = re.sub(r"\n```$", "", json_str)
        return json_str.strip()

        # GPT-4o 调用

    def extract_data_from_image(image_path_list):
        # 收集转换格式后的image地址
        image_path_base64_list = []
        for image_index, image_path in enumerate(image_path_list):
            response = requests.get(image_path, verify=False)
            if response.status_code == 200:
                image_bytes = response.content
                base64_img = base64.b64encode(image_bytes).decode('utf-8')
                image_path_base64_list.append(base64_img)
            else:
                raise Exception(f"Failed to fetch image {image_index}. Status code: {response.status_code}")

        # 提取字段 Prompt 模板
        PROMPT_TEMPLATE = """
        你将看到多张由pdf转成的图片，是按照页码顺序传输的，这些图片中可能包含一张或多张产品电器架构子图（electric architecture diagram）。请你分析每张图，并按如下要求提取结构信息：

        ---

        📌 **一、判断每张电器架构子图是否包含以下“已知组件”（known modules）：**

        以下是需识别的组件及**识别规则说明**：

        1. **MCU**：图中显示为 Power Inverter 组件。
        2. **DCAC**：图中标注为 DC/AC。
        3. **DCDC**：图中标注为 DC/DC（如 12V DC/DC）。
        4. **OBC**：图中标注为 OBC。
        5. **(K3/K4)**：图上有 K3 和 K4 两个开关。
        6. **(K1/K2)**：图上有 K1 和 K2 两个开关。
        7. **PDU**：组件名为 PDU。
        8. **HVAC**：组件名为 HVAC。
        9. **PTC**：组件名为 PTC。
        10. **(K5)**：图上有 K5 开关。
        11. **PTC2**：组件名可能为PTC2 或者 OBC 或者就是 Batt. Heating 也可能是这几种的组合。
        12. **(K6)**：图上有 K6 开关。
        13. **PTO1**：组件名为 PTO1。
        14. **PTO2**：组件名为 PTO2。
        15. **(K8/K9)**：图上有 K8 和 K9 两个开关。

        ---

        📌 **二、同时识别图中存在但不属于上述列表的其他组件，归为 `unkown_modules`。**

        ---

        📌 **三、请提取该架构图的图名或表名作为键名（如标题“XIN1 electric architecture”），一般位置在图片的下方。**

        ---

        📌 **四、返回格式必须为以下标准 JSON 格式**（注意结构），每张图的信息为返回值中的一组键值对：

        ```json
        {
          "XIN1 electric architecture": {
            "known_modules": [
              "MCU",
              "Fast Charge(K3/K4)",
              ...
            ],
            "unkown_modules": [
              "XYZ-123",
              ...
            ]
          }
        }



        如果一页有          
       ⚠注意事项：
       - 识别到已知组件进行名称存储时一定要和上面列举的已知组件名称保持统一，有括号的不要少括号，大小写也保持一致；
       - 如果图片内容中没有电器架构子图，即使出现了组件名称，也请返回空的JSON；
       - 严格遵守组件映射规则：
            - Power Inverter → MCU
       - 图片的名称和图片本身可能被分页面分割所以处在不同图片中，请注意页码来判别图片的名称和图片的关系。
       ⚠返回结果必须是 JSON 格式，不要添加任何解释说明、注释或额外内容。
       """

        # 喂给大模型需要一次读取的图片
        response_content_list = [{"type": "text", "text": PROMPT_TEMPLATE}]
        for image_path_base64 in image_path_base64_list:
            response_content_image_dict = {}
            response_content_image_dict['type'] = "image_url"
            response_content_image_dict['image_url'] = {
                "url": f"data:image/png;base64,{image_path_base64}"
            }
            response_content_list.append(response_content_image_dict)

        response = azure_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "你是一个图像信息提取专家，擅长从截图中读取结构化信息。"
                },
                {
                    "role": "user",
                    "content": response_content_list
                }
            ],
            temperature=0
        )
        return response.choices[0].message.content

    pdf_page_related_to_eletric_architecture_image_path_list = []
    doc = fitz.open(pdf_path)
    # 定位到本页 考虑 可能会分页向下兼容一页
    for page_number in [start_page, start_page + 1]:
        print(f"convert pdf page number : {page_number} in  image for LLM read")
        page = doc[page_number]
        # 转image
        pix = page.get_pixmap(dpi=200)
        current_file_path = os.path.abspath(__file__)
        # 获取当前文件的根目录
        root_directory = os.path.dirname(current_file_path)
        image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{page_number - 2}.png"
        # docker 存储 image
        local_img_path = os.path.join(root_directory, 'tcd_pdf_images', f"{image_filename}")
        print("local_img_path: ", local_img_path)
        pix.save(local_img_path)
        with open(local_img_path, "rb") as image:
            files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
            upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
            response = requests.post(upload_url, files=files, verify=False)
            image_path = response.json()['data'][0]['url']
            print("document_server_path: ", image_path)
            pdf_page_related_to_eletric_architecture_image_path_list.append(image_path)
    try:
        # 同时解析两张图
        json_str = extract_data_from_image(pdf_page_related_to_eletric_architecture_image_path_list)

        print("🔍 LLM 原始返回内容：", repr(json_str))

        json_str_clean = clean_llm_json(json_str)
        print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容

        if json_str_clean.strip():  # 防止空字符串
            data_list = json.loads(json_str_clean)
            for key, value in data_list.items():
                known_modules_list = value["known_modules"]

                if '(K1/K2)' in known_modules_list:
                    print("(K1/K2) CHECK")
                    known_modules_list.remove('(K1/K2)')
                    known_modules_list.append('PDU_Main Circuit(K1/K2)')

                if '(K3/K4)' in known_modules_list:
                    print("(K3/K4) CHECK")
                    known_modules_list.remove('(K3/K4)')
                    known_modules_list.append('PDU_Fast Charge(K3/K4)')

                # # check ptc & k5 是否同时存在 合并成一组
                if 'PTC' in known_modules_list and '(K5)' in known_modules_list:
                    print("PTC (K5) CHECK")
                    known_modules_list.remove('PTC')
                    known_modules_list.remove('(K5)')
                    known_modules_list.append('PDU_PTC(K5)')

                if 'PTO2' in known_modules_list and '(K8/K9)' in known_modules_list:
                    print("PTO2 (K8/K9) CHECK")
                    known_modules_list.remove('PTO2')
                    known_modules_list.append('PDU_PTO2(K8/K9)')

                if 'PTO1' in known_modules_list and '(K8/K9)' in known_modules_list:
                    print("PTO1 (K8/K9) CHECK")
                    known_modules_list.remove('PTO1')
                    known_modules_list.remove('(K8/K9)')
                    known_modules_list.append('PDU_PTO1(K8/K9)')

                if 'PTC2' in known_modules_list and '(K6)' in known_modules_list:
                    print("PTC2 (K6) CHECK")
                    known_modules_list.remove('PTC2')
                    known_modules_list.remove('(K6)')
                    known_modules_list.append('PDU_PTC2(K6)')

                if 'PDU' in known_modules_list:
                    print(1)
                    known_modules_list.remove('PDU')
                if 'HVAC' in known_modules_list:
                    print(4)
                    known_modules_list.remove('HVAC')
                    known_modules_list.append("PDU_HVAC")
                if 'PTC' in known_modules_list:
                    print(5)
                    known_modules_list.remove('PTC')
                    known_modules_list.append("PDU_PTC")

                if 'PTC2' in known_modules_list:
                    print(7)
                    known_modules_list.remove('PTC2')
                    known_modules_list.append("PDU_PTC2")
                if 'PTO2' in known_modules_list:
                    print(10)
                    known_modules_list.remove('PTO2')
                    known_modules_list.append("PDU_PTO2)")


            print("返回的prv表格: ", data_list)

            # 异步传输每个image和content
            result_dict = {'page_image_path': pdf_page_related_to_eletric_architecture_image_path_list,
                           'page_image_content': data_list}
            print("！！！ 发送给前端的结果： ", result_dict)
            prv_message = {'connectionID': task_id,
                           'category': 'tcd_electric_architecture_image', 'from': '', 'to': '',
                           'message': json.dumps(result_dict, ensure_ascii=False),
                           'remarks': json.dumps({'paragraph_start': 1,
                                                  'response_end': 0})}
            try:
                asyncio.run(websocket_client(prv_message))
            except:
                time.sleep(0.5)
                asyncio.run(websocket_client(prv_message))

    except Exception as e:
        print(f"❌ Error parsing {pdf_page_related_to_eletric_architecture_image_path_list}: {e}")


    # 异步传输结束信号空字符串
    prv_message = {'connectionID': task_id,
                   'category': 'text', 'from': '', 'to': '',
                   'message': '',
                   'remarks': json.dumps({'paragraph_start': 0,
                                          'response_end': 1})}
    try:
        asyncio.run(websocket_client(prv_message))
    except:
        time.sleep(0.5)
        asyncio.run(websocket_client(prv_message))


#######################################################################################################

                              #  TCD INTERFACE TABLE  #

#######################################################################################################
def find_interface_table_page(pdf_path):
    start_keyword = "3.2.2.2 HV INTERFACE"
    end_keyword = "3.2.3.2 LV SIGNAL INTERFACE"
    start_page = None
    end_page = None
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if start_keyword in text and start_page is None:
                print('find start page')
                start_page = i
            elif end_keyword in text and start_page is not None:
                print('find end page')
                end_page = i
                break
    return start_page, end_page


def auto_tcd_interface_table_extract(task_id, pdf_path, start_page, end_page):
    def clean_llm_json(json_str):
        # 去掉前后的 markdown 代码块标识符 ```json ... ```
        if json_str.startswith("```"):
            json_str = re.sub(r"^```(?:json)?\n", "", json_str)
            json_str = re.sub(r"\n```$", "", json_str)
        return json_str.strip()

        # GPT-4o 调用

    def extract_data_from_image(image_path_list):
        # 收集转换格式后的image地址
        image_path_base64_list = []
        for image_index, image_path in enumerate(image_path_list):
            response = requests.get(image_path, verify=False)
            if response.status_code == 200:
                image_bytes = response.content
                base64_img = base64.b64encode(image_bytes).decode('utf-8')
                image_path_base64_list.append(base64_img)
            else:
                raise Exception(f"Failed to fetch image {image_index}. Status code: {response.status_code}")
        # 提取字段 Prompt 模板
        PROMPT_TEMPLATE = """
        你将看到多张由pdf转成的图片，是按照页码顺序传输的，这些图片中可能包含多个组件接口的表格信息，请读取除了Battery和LV SIGNAL的所有接口表格信息。
        需要注意一张完整的表格可能被分页面分割所以处在不同图片中，请智能识别进行拼接，表格信息包含以下几列：
        - 项目 Item
        - 最小值/Min
        - 最大值/Max
        - 单位/Unit
        - 备注 Comments

        请你识别所有组件接口相关表格，提取表格名称和表格内容，返回格式必须为以下标准 JSON 格式**（注意结构），
        每张表格的信息为返回值中的一组键值对，键为表格名称，值为list套list，有多少行就套多少组list，第一组是列名：

         ```json
        {
         "HV PHASE INTERFACE":[
            ["项目 Item","最小值/Min","最大值/Max","单位/Unit","备注 Comments"],
            ["高压相电流（持续时间小于60 秒）HV phase current at UVW, t < 60s","-380","380","Arms","双向电流（驱动模式/制动模式）Current flow in both directions (due to motor / generator- mode) @MAX.Tcool=65℃, Q=10L/min"],
            ["高压相电流（持续）","-200","200","Arms","双向电流（驱动模式/制动模式"],
            ...
            ],
        "DC CHARGE INTERFACE":[
            ["项目 Item","最小值/Min","最大值/Max","单位/Unit","备注 Comments"],
            [],
            [],
            ...
        ]
            
        }

       ⚠注意事项：
       - 请一定只提取接口相关的表格信息；
       - 若某些字段为空，请也保留字段并赋值为空字符串 ""，不要串行。
       - 对于表格尾出现的整行注意等额外标注不要返回。
       ⚠返回结果必须是标准合法的 JSON 格式，不要添加任何解释说明、注释或额外内容。
       """

        # 喂给大模型需要一次读取的图片
        response_content_list = [{"type": "text", "text": PROMPT_TEMPLATE}]
        for image_path_base64 in image_path_base64_list:
            response_content_image_dict = {}
            response_content_image_dict['type'] = "image_url"
            response_content_image_dict['image_url'] = {
                "url": f"data:image/png;base64,{image_path_base64}"
            }
            response_content_list.append(response_content_image_dict)

        response = azure_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "你是一个图像信息提取专家，擅长从图中读取结构化信息。"
                },
                {
                    "role": "user",
                    "content": response_content_list
                }
            ],
            temperature=0
        )
        return response.choices[0].message.content

    pdf_page_related_to_interface_table_path_list = []
    doc = fitz.open(pdf_path)
    # 读取定位的page内容
    for page_number in range(start_page, end_page + 1):
        print(f"convert pdf page number : {page_number} in  image for LLM read")
        page = doc[page_number]
        # 转image
        pix = page.get_pixmap(dpi=200)
        current_file_path = os.path.abspath(__file__)
        # 获取当前文件的根目录
        root_directory = os.path.dirname(current_file_path)
        image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{page_number - 2}.png"
        # docker 存储 image
        local_img_path = os.path.join(root_directory, 'tcd_pdf_images', f"{image_filename}")
        print("local_img_path: ", local_img_path)
        pix.save(local_img_path)
        with open(local_img_path, "rb") as image:
            files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
            upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
            response = requests.post(upload_url, files=files, verify=False)
            image_path = response.json()['data'][0]['url']
            print("document_server_path: ", image_path)
            pdf_page_related_to_interface_table_path_list.append(image_path)

    try:
        json_str = extract_data_from_image(pdf_page_related_to_interface_table_path_list)
        print("🔍 LLM 原始返回内容：", repr(json_str))

        json_str_clean = clean_llm_json(json_str)
        print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容

        if json_str_clean.strip():  # 防止空字符串
            data_list = json.loads(json_str_clean)
            print("返回的prv表格: ", data_list)

            # 异步传输每个image和content
            result_dict = {'page_image_path': pdf_page_related_to_interface_table_path_list,
                           'page_image_content': data_list}
            print("！！！ 发送给前端的结果： ", result_dict)
            prv_message = {'connectionID': task_id,
                           'category': 'tcd_interface_table', 'from': '', 'to': '',
                           'message': json.dumps(result_dict, ensure_ascii=False),
                           'remarks': json.dumps({'paragraph_start': 1,
                                                  'response_end': 0})}
            try:
                asyncio.run(websocket_client(prv_message))
            except:
                time.sleep(0.5)
                asyncio.run(websocket_client(prv_message))

    except Exception as e:
        print(f"❌ Error parsing {pdf_page_related_to_interface_table_path_list}: {e}")

    # 异步传输结束信号空字符串
    prv_message = {'connectionID': task_id,
                   'category': 'text', 'from': '', 'to': '',
                   'message': '',
                   'remarks': json.dumps({'paragraph_start': 0,
                                          'response_end': 1})}
    try:
        asyncio.run(websocket_client(prv_message))
    except:
        time.sleep(0.5)
        asyncio.run(websocket_client(prv_message))


#######################################################################################################

                              #  TCD CHARACTERISTICS TABLE  #

#######################################################################################################

def find_char_table_page(pdf_path):
    start_keyword = "4.1.3.1 ELECTRICAL CHARACTERISTICS"
    end_keyword = "4.2.1.1 APPROPRIATE USE"
    start_page = None
    end_page = None
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if start_keyword in text and start_page is None:
                print('find start page')
                start_page = i
            elif end_keyword in text and start_page is not None:
                print('find end page')
                end_page = i
                break
    return start_page, end_page


def auto_tcd_char_table_extract(task_id, pdf_path, start_page, end_page):
    def clean_llm_json(json_str):
        # 去掉前后的 markdown 代码块标识符 ```json ... ```
        if json_str.startswith("```"):
            json_str = re.sub(r"^```(?:json)?\n", "", json_str)
            json_str = re.sub(r"\n```$", "", json_str)
        return json_str.strip()

        # GPT-4o 调用

    def extract_data_from_image(image_path_list):
        # 收集转换格式后的image地址
        image_path_base64_list = []
        for image_index, image_path in enumerate(image_path_list):
            response = requests.get(image_path, verify=False)
            if response.status_code == 200:
                image_bytes = response.content
                base64_img = base64.b64encode(image_bytes).decode('utf-8')
                image_path_base64_list.append(base64_img)
            else:
                raise Exception(f"Failed to fetch image {image_index}. Status code: {response.status_code}")
        # 提取字段 Prompt 模板
        PROMPT_TEMPLATE = """
        你将看到多张由pdf转成的图片，是按照页码顺序传输的，这些图片中可能包含逆变器(INVERTER)、PDU、DCDC、DCAC这些组件特性的表格信息。

        请你识别所有相关特性表格，提取表格名称和表格内容，每张表格的信息为返回值中的一组键值对，
        键为表格名称，值为list套list，有多少行就套多少组list，第一组一定是列名，返回格式必须为以下标准 JSON 格式**（注意结构）：

         ```json
        {
         "DCAC inputs and outputs characteristics/ DCAC输入输出特性":[
            ["项目 Item","最小值 Min","最大值 Max","单位 Unit","备注 Comments"],
            ["高压直流输入 HV DC Input","","","",""],
            ["逆变器工作范围 Operation of Inverter without limitation","350","700","V","@ Tcool_max =65℃, Q =10L/min"],
            ["UVW (相线) UVW(Phase line)","","","",""],
            ["输出相电流能力 Output Phase Current","","25","Arms","@ Tcool_max =65℃, Q =10L/min"],
            ...
            ],
        "DCDC inputs and outputs characteristics/ DCDC输入输出特性":[
            ["项目 Item","最小值/Min","最大值/Max","单位/Unit","备注 Comments"],
            [],
            [],
            ...
        ]


        }


       ⚠注意事项：
       - 请一定只提取逆变器(INVERTER)、PDU、DCDC、DCAC这些组件特性相关的表格信息，其它特性表格不要提取；
       - 如果遇到被页面切开的表格，请一定不要串列串行，若某些格子为空，请返回空字符串""表示；
       - 请一定看清表格的边框，不要出现非空格子和空格子内容进行位置交换的情况；
       - 当读取的每行的第二个和第三个位置,其中一个是空字符串""另一个是非空字符串，请确保非空字符串在第三个位置；
       - 若某些格子为空，请也保留赋值为空字符串 ""，不要串行。
       ⚠返回结果必须是标准合法的 JSON 格式，不要添加任何解释说明、注释或额外内容。
       """

        # 喂给大模型要一次读取的图片
        response_content_list = [{"type": "text", "text": PROMPT_TEMPLATE}]
        for image_path_base64 in image_path_base64_list:
            response_content_image_dict = {}
            response_content_image_dict['type'] = "image_url"
            response_content_image_dict['image_url'] = {
                "url": f"data:image/png;base64,{image_path_base64}"
            }
            response_content_list.append(response_content_image_dict)

        response = azure_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "你是一个图像信息提取专家，擅长从截图中读取结构化信息。"
                },
                {
                    "role": "user",
                    "content": response_content_list
                }
            ],
            temperature=0
        )
        return response.choices[0].message.content

    pdf_page_related_to_char_table_path_list = []
    doc = fitz.open(pdf_path)
    # 读取定位的page内容
    for page_number in range(start_page, end_page + 1):
        print(f"convert pdf page number : {page_number} in  image for LLM read")
        page = doc[page_number]
        # 转image
        pix = page.get_pixmap(dpi=200)
        current_file_path = os.path.abspath(__file__)
        # 获取当前文件的根目录
        root_directory = os.path.dirname(current_file_path)
        image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{page_number - 2}.png"
        # docker 存储 image
        local_img_path = os.path.join(root_directory, 'tcd_pdf_images', f"{image_filename}")
        print("local_img_path: ", local_img_path)
        pix.save(local_img_path)
        with open(local_img_path, "rb") as image:
            files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
            upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
            response = requests.post(upload_url, files=files, verify=False)
            image_path = response.json()['data'][0]['url']
            print("document_server_path: ", image_path)
            pdf_page_related_to_char_table_path_list.append(image_path)

    try:
        json_str = extract_data_from_image(pdf_page_related_to_char_table_path_list)
        print("🔍 LLM 原始返回内容：", repr(json_str))

        json_str_clean = clean_llm_json(json_str)
        print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容

        if json_str_clean.strip():  # 防止空字符串
            data_list = json.loads(json_str_clean)
            print("返回的prv表格: ", data_list)

            # 异步传输每个image和content
            result_dict = {'page_image_path': pdf_page_related_to_char_table_path_list,
                           'page_image_content': data_list}
            print("！！！ 发送给前端的结果： ", result_dict)
            prv_message = {'connectionID': task_id,
                           'category': 'tcd_characteristics_table', 'from': '', 'to': '',
                           'message': json.dumps(result_dict, ensure_ascii=False),
                           'remarks': json.dumps({'paragraph_start': 1,
                                                  'response_end': 0})}
            try:
                asyncio.run(websocket_client(prv_message))
            except:
                time.sleep(0.5)
                asyncio.run(websocket_client(prv_message))

    except Exception as e:
        print(f"❌ Error parsing {pdf_page_related_to_char_table_path_list}: {e}")

    # 异步传输结束信号空字符串
    prv_message = {'connectionID': task_id,
                   'category': 'text', 'from': '', 'to': '',
                   'message': '',
                   'remarks': json.dumps({'paragraph_start': 0,
                                          'response_end': 1})}
    try:
        asyncio.run(websocket_client(prv_message))
    except:
        time.sleep(0.5)
        asyncio.run(websocket_client(prv_message))



#######################################################################################################

                              #  TCD PN LIST  #

#######################################################################################################


def auto_tcd_pn_table_extract(pdf_path, page_number=3):
    def clean_llm_json(json_str):
        # 去掉前后的 markdown 代码块标识符 ```json ... ```
        if json_str.startswith("```"):
            json_str = re.sub(r"^```(?:json)?\n", "", json_str)
            json_str = re.sub(r"\n```$", "", json_str)
        return json_str.strip()

        # GPT-4o 调用

    def extract_data_from_image(image_path):
        response = requests.get(image_path, verify=False)
        if response.status_code == 200:
            image_bytes = response.content
            base64_img = base64.b64encode(image_bytes).decode('utf-8')
        else:
            raise Exception(f"Failed to fetch image. Status code: {response.status_code}")
        # 提取字段 Prompt 模板
        PROMPT_TEMPLATE = """
               你将看到一张由pdf转成的图片，图片中可能包含一个或多个开发零件号(part number)。请提取所有出现的开发零件号(part number)，注意不要和供货图纸号(offer drawing)混淆,
               返回格式必须为以下标准 JSON 格式**（注意结构），一组键值对，part_numbe为键，值为list包含所有开发零件号(part number)：
               ```json
               {"part_number":
                    ["0437CX001F",
                       "0437CX001G",
                       "0437CX001H",
                       ...
                    ]
               }
               ⚠注意事项：
               - 请一定只提取开发零件号(part number)相关的表格信息；
               ⚠返回结果必须是标准合法的 JSON 格式，不要添加任何解释说明、注释或额外内容。
               """
        response = azure_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "你是一个图像信息提取专家，擅长从截图中读取结构化信息。"
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": PROMPT_TEMPLATE},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{base64_img}"
                            }
                        }
                    ]
                }
            ],
            temperature=0
        )
        return response.choices[0].message.content

    doc = fitz.open(pdf_path)
    # 定位到内容开始的第一页 包含料号信息
    print(f"convert pdf page number : {page_number} in  image for LLM read")
    page = doc[page_number]
    # 转image
    pix = page.get_pixmap(dpi=200)
    current_file_path = os.path.abspath(__file__)
    # 获取当前文件的根目录
    root_directory = os.path.dirname(current_file_path)
    image_filename = f"{os.path.basename(pdf_path).replace('.pdf', '')}_page{page_number - 2}.png"
    # docker 存储 image
    local_img_path = os.path.join(root_directory, 'tcd_pdf_images', f"{image_filename}")
    print("local_img_path: ", local_img_path)
    pix.save(local_img_path)
    with open(local_img_path, "rb") as image:
        files = {"file": (image_filename, image, "image/png")}  # 定义文件信息
        upload_url = 'https://szhlinvma75.apac.bosch.com:59108/api/upload'
        response = requests.post(upload_url, files=files, verify=False)
        image_path = response.json()['data'][0]['url']
        try:
            json_str = extract_data_from_image(image_path)
            print("🔍 LLM 原始返回内容：", repr(json_str))

            json_str_clean = clean_llm_json(json_str)
            print("🧹 处理后的内容：", repr(json_str_clean))  # 注意用 repr 打印确保看到真实内容

            if json_str_clean.strip():  # 防止空字符串
                data_list = json.loads(json_str_clean)
                print("返回的prv表格: ", data_list)
                data_list['part_number'] = [s.replace('.', '').upper() for s in data_list['part_number']]
                print("返回的prv表格，处理后的标准pn: ", data_list)



        except Exception as e:
            print(f"❌ Error parsing {image_path}: {e}")
            data_list = {'part_number': []}

    return data_list






if __name__ == "__main__":
    # digital age 2025
    result = get_agent_reply("AGENT, COST,simutanoust，people， RAG。")
    print(json.dumps(result, ensure_ascii=False))
    # auto prv 示例用法
    # result = auto_prv_improve("123",r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE_WUJ\Testing program auto generation\input_pdf\91_PDFsam_GeeA_2.0-8.pdf')
    # print(json.dumps(result, ensure_ascii=False))

    #auto_prv_improve_by_action 示例用法
    # result = auto_prv_improve_by_action("123",r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE_WUJ\second_phase\splitpdf\41_PDFsam_Geely_2.0_C.pdf', r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE_WUJ\second_phase\Geely2.0_FCT50-Additional file.xlsx')
    # print(json.dumps(result, ensure_ascii=False))


    # tcd electric architecture 示例用法
    # pdf_path = r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE\文件\TCD\E820\TCD_Xin1_WeiChai_Export_v1.pdf'
    # # # target_title = "3.2.2 ELECTRICAL INTERFACE"
    # page_number = find_electric_architecture_page(pdf_path)
    # if page_number:
    #     print(f"第二次出现位于第 {page_number} 页")
    # else:
    #     print("未找到第二次出现")
    # auto_tcd_electric_architecture_extract("123", pdf_path, page_number)

    # tcd interface table 示例用法
    # pdf_path = r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE\文件\TCD\E820\TCD_Xin1_JMC_E820_0437CD0002_v1.pdf'
    # start_page, end_page = find_interface_table_page(pdf_path)
    # print(start_page, end_page)
    # auto_tcd_interface_table_extract("123", pdf_path, start_page, end_page)

    # tcd char table示例用法
    # pdf_path = r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE\文件\TCD\E820\TCD_Xin1_JMC_E820_v1.pdf'
    # start_page, end_page = find_char_table_page(pdf_path)
    # print(start_page, end_page)
    # auto_tcd_char_table_extract("123", pdf_path, start_page, end_page)

    # pn_info示例用法
    # pdf_path = r'C:\Users\GNW1SZH\OneDrive - Bosch Group\PersonalDrive\项目\2025\digitalSE\文件\TCD\E820\TCD_Xin1_JMC_E820_v1.pdf'
    # auto_tcd_pn_table_extract(pdf_path)
